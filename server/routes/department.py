import io
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone

from flask import Blueprint, request, jsonify, make_response
from sqlalchemy import case, func
from models import db, User, Section, Batch, BatchSection, Faculty, FacultyBatchSection, Student, Department, College, AttendanceSession, AttendanceRecord, Subject, Semester, Program, Mark, TimetableEntry, SectionSubjectFormat, SectionSubjectAssignment, FormatSubject
from auth import hash_password, normalize_email, token_required, role_required

dept_bp = Blueprint('department', __name__, url_prefix='/api/department')

TIMETABLE_DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
TIMETABLE_SLOTS = [
    '09:00-10:40',   # slot 0  (100 min)
    '10:40-12:20',   # slot 1  (100 min)
    '01:30-03:10',   # slot 2  (100 min)
    '03:10-04:00',   # slot 3  (50 min)
]

NON_TEACHING_LABELS = [
    'Library',
    'Tutorial',
    'Mentoring',
    'Sports',
    'Placement Prep',
    'Project Work',
    'Seminar',
    'Club Activity',
]

DEFAULT_SUBJECT_TYPES = ['Common', 'Elective', 'Open Elective']


def normalize_subject_type(subject_type):
    value = (subject_type or '').strip()
    if not value:
        return ''
    lowered = value.lower()
    if lowered == 'common':
        return 'Common'
    if lowered == 'elective':
        return 'Elective'
    if lowered in ('open elective', 'open-elective', 'open_elective'):
        return 'Open Elective'
    return value


def get_dept_subject_prefix(dept_name):
    tokens = re.findall(r'[A-Za-z]+', (dept_name or '').upper())
    if not tokens:
        return 'DP'
    base = tokens[0]
    # CSE -> CS, ECE -> EC, EEE -> EE
    if len(base) >= 3 and base.endswith('E'):
        base = base[:-1]
    if len(base) < 2:
        base = ''.join(tokens)
    return (base[:3] or 'DP')


def semester_to_year_sem(semester_no):
    year_no = ((semester_no - 1) // 2) + 1
    sem_in_year = 1 if semester_no % 2 == 1 else 2
    return year_no, sem_in_year


def subject_code_sort_key(code):
    code_value = (code or '').upper()
    match = re.match(r'^([A-Z]+)(\d+)$', code_value)
    if not match:
        return (code_value, 0)
    prefix, number = match.groups()
    return (prefix, int(number))


def generate_subject_code_for_dept(dept_name, semester_no):
    prefix = get_dept_subject_prefix(dept_name)
    year_no, sem_in_year = semester_to_year_sem(semester_no)
    code_base = f'{prefix}{year_no}{sem_in_year}'
    matcher = re.compile(rf'^{re.escape(code_base)}(\d{{2}})$')
    max_serial = 0
    for (existing_code,) in db.session.query(Subject.subject_code).filter(Subject.subject_code.like(f'{code_base}%')).all():
        match = matcher.match(existing_code or '')
        if match:
            max_serial = max(max_serial, int(match.group(1)))
    next_serial = max_serial + 1
    while True:
        candidate = f'{code_base}{next_serial:02d}'
        if not db.session.get(Subject, candidate):
            return candidate
        next_serial += 1


def clear_batch_subject_setup(batch_id, section_ids=None):
    format_ids_query = db.session.query(SectionSubjectFormat.id).filter_by(batch_id=batch_id)
    format_ids = [row[0] for row in format_ids_query.all()]
    if format_ids:
        FormatSubject.query.filter(FormatSubject.format_id.in_(format_ids)).delete(synchronize_session=False)
        SectionSubjectFormat.query.filter_by(batch_id=batch_id).delete(synchronize_session=False)

    assignment_query = SectionSubjectAssignment.query
    if section_ids:
        assignment_query = assignment_query.filter(SectionSubjectAssignment.section_id.in_(section_ids))
    else:
        batch_section_ids = [row[0] for row in db.session.query(Section.section_id).filter_by(batch_id=batch_id).all()]
        if batch_section_ids:
            assignment_query = assignment_query.filter(SectionSubjectAssignment.section_id.in_(batch_section_ids))
        else:
            assignment_query = assignment_query.filter(False)
    assignment_query.delete(synchronize_session=False)

    TimetableEntry.query.filter_by(batch_id=batch_id).delete(synchronize_session=False)



def escape_pdf_text(value):
    return (
        str(value)
        .replace('\\', '\\\\')
        .replace('(', '\\(')
        .replace(')', '\\)')
    )


def build_pdf_document(page_stream, page_width=842, page_height=595):
    pdf = bytearray(b'%PDF-1.4\n')
    offsets = []

    def add_object(obj_id, content_bytes):
        offsets.append(len(pdf))
        pdf.extend(f'{obj_id} 0 obj\n'.encode('latin-1'))
        pdf.extend(content_bytes)
        pdf.extend(b'\nendobj\n')

    add_object(1, b'<< /Type /Catalog /Pages 2 0 R >>')
    add_object(2, b'<< /Type /Pages /Count 1 /Kids [5 0 R] >>')
    add_object(3, b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>')
    add_object(4, b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>')
    add_object(
        5,
        (
            f'<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {page_width} {page_height}] '
            f'/Resources << /Font << /F1 3 0 R /F2 4 0 R >> >> /Contents 6 0 R >>'
        ).encode('latin-1')
    )
    stream = page_stream.encode('latin-1', errors='replace')
    stream_object = b'<< /Length ' + str(len(stream)).encode('latin-1') + b' >>\nstream\n' + stream + b'\nendstream'
    add_object(6, stream_object)

    xref_offset = len(pdf)
    pdf.extend(f'xref\n0 {len(offsets) + 1}\n'.encode('latin-1'))
    pdf.extend(b'0000000000 65535 f \n')
    for offset in offsets:
        pdf.extend(f'{offset:010d} 00000 n \n'.encode('latin-1'))
    pdf.extend(
        (
            f'trailer\n<< /Size {len(offsets) + 1} /Root 1 0 R >>\n'
            f'startxref\n{xref_offset}\n%%EOF'
        ).encode('latin-1')
    )
    return bytes(pdf)


def build_pdf_pages(page_streams, page_width=842, page_height=595):
    pdf = bytearray(b'%PDF-1.4\n')
    offsets = []
    page_count = len(page_streams)
    pages_object_id = 2
    font_regular_id = 3
    font_bold_id = 4
    first_page_id = 5

    def add_object(obj_id, content_bytes):
        offsets.append((obj_id, len(pdf)))
        pdf.extend(f'{obj_id} 0 obj\n'.encode('latin-1'))
        pdf.extend(content_bytes)
        pdf.extend(b'\nendobj\n')

    add_object(1, b'<< /Type /Catalog /Pages 2 0 R >>')
    page_refs = [f'{first_page_id + (index * 2)} 0 R' for index in range(page_count)]
    add_object(pages_object_id, f'<< /Type /Pages /Count {page_count} /Kids [{" ".join(page_refs)}] >>'.encode('latin-1'))
    add_object(font_regular_id, b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>')
    add_object(font_bold_id, b'<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>')

    for index, stream_text in enumerate(page_streams):
        page_id = first_page_id + (index * 2)
        content_id = page_id + 1
        add_object(
            page_id,
            (
                f'<< /Type /Page /Parent {pages_object_id} 0 R /MediaBox [0 0 {page_width} {page_height}] '
                f'/Resources << /Font << /F1 {font_regular_id} 0 R /F2 {font_bold_id} 0 R >> >> /Contents {content_id} 0 R >>'
            ).encode('latin-1')
        )
        stream = stream_text.encode('latin-1', errors='replace')
        add_object(content_id, b'<< /Length ' + str(len(stream)).encode('latin-1') + b' >>\nstream\n' + stream + b'\nendstream')

    object_count = max(obj_id for obj_id, _ in offsets)
    offset_map = {obj_id: offset for obj_id, offset in offsets}
    xref_offset = len(pdf)
    pdf.extend(f'xref\n0 {object_count + 1}\n'.encode('latin-1'))
    pdf.extend(b'0000000000 65535 f \n')
    for obj_id in range(1, object_count + 1):
        pdf.extend(f'{offset_map[obj_id]:010d} 00000 n \n'.encode('latin-1'))
    pdf.extend(
        (
            f'trailer\n<< /Size {object_count + 1} /Root 1 0 R >>\n'
            f'startxref\n{xref_offset}\n%%EOF'
        ).encode('latin-1')
    )
    return bytes(pdf)


def build_pdf_text(x, y, text, font='F1', size=10):
    escaped = escape_pdf_text(text)
    return f'0 g 0 G BT /{font} {size} Tf 1 0 0 1 {x:.2f} {y:.2f} Tm ({escaped}) Tj ET'


def build_pdf_centered_text(x, y, width, text, font='F1', size=10):
    approx_width = len(str(text)) * size * 0.5
    start_x = x + max(0, (width - approx_width) / 2)
    return build_pdf_text(start_x, y, text, font=font, size=size)


def build_pdf_rect(x, y, width, height, fill_gray=None, stroke_gray=0):
    commands = []
    if fill_gray is not None:
        commands.append(f'{fill_gray:.2f} g {x:.2f} {y:.2f} {width:.2f} {height:.2f} re f')
    commands.append(f'{stroke_gray:.2f} G {x:.2f} {y:.2f} {width:.2f} {height:.2f} re S')
    return '\n'.join(commands)


def truncate_text(text, max_chars):
    value = str(text)
    return value if len(value) <= max_chars else value[:max_chars - 3] + '...'


def number_to_words(value):
    ones = [
        'zero', 'one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine',
        'ten', 'eleven', 'twelve', 'thirteen', 'fourteen', 'fifteen', 'sixteen',
        'seventeen', 'eighteen', 'nineteen',
    ]
    tens = ['', '', 'twenty', 'thirty', 'forty', 'fifty', 'sixty', 'seventy', 'eighty', 'ninety']

    def integer_words(number):
        number = int(number)
        if number < 20:
            return ones[number]
        if number < 100:
            return tens[number // 10] if number % 10 == 0 else f'{tens[number // 10]} {ones[number % 10]}'
        if number < 1000:
            suffix = integer_words(number % 100) if number % 100 else ''
            return f'{ones[number // 100]} hundred{f" {suffix}" if suffix else ""}'
        return str(number)

    numeric = float(value or 0)
    if numeric.is_integer():
        return integer_words(int(numeric)).title()
    whole = int(numeric)
    decimal = str(numeric).split('.', 1)[1].rstrip('0')
    decimal_words = ' '.join(ones[int(digit)] for digit in decimal)
    return f'{integer_words(whole)} point {decimal_words}'.title()


def build_timetable_pdf(section, entries):
    department = db.session.get(Department, section.dept_id)
    college = db.session.get(College, department.college_id) if department and department.college_id else None
    batch = db.session.get(Batch, section.batch_id)
    entry_map = {(entry.day_order, entry.slot_index): entry for entry in entries}
    unique_subjects = []
    seen_subjects = set()
    for entry in entries:
        if entry.subject_code in seen_subjects:
            continue
        seen_subjects.add(entry.subject_code)
        unique_subjects.append((entry.subject_code, entry.subject_name or '-', entry.faculty_name or '-'))

    page_width = 842
    page_height = 595
    margin = 28
    content_width = page_width - (margin * 2)
    top_y = page_height - margin
    commands = ['1.10 w', '0 g', '0 G']

    commands.append(build_pdf_rect(margin - 6, margin - 6, content_width + 12, page_height - (margin * 2) + 12))
    commands.append(build_pdf_rect(margin, top_y - 24, content_width, 18, fill_gray=0.80))
    commands.append(build_pdf_centered_text(margin, top_y - 18, content_width, f'DEPARTMENT OF {department.dept_name.upper() if department else "ACADEMICS"}', font='F2', size=11))
    commands.append(build_pdf_centered_text(margin, top_y - 38, content_width,'ANDHRA UNIVERSITY', font='F2', size=11))
    meta_top = top_y - 78
    meta_height = 34
    meta_cols = [content_width * 0.38, content_width * 0.18, content_width * 0.18, content_width * 0.26]
    meta_labels = [
        f'BATCH: {batch.batch_name if batch else "-"}',
        f'SEM: {section.current_semester}',
        f'SEC: {section.section_name}',
        f'DEPT: {department.dept_name if department else "-"}',
    ]
    current_x = margin
    for width, label in zip(meta_cols, meta_labels):
        commands.append(build_pdf_rect(current_x, meta_top - meta_height, width, meta_height))
        commands.append(build_pdf_centered_text(current_x, meta_top - 21, width, label, font='F2', size=9))
        current_x += width

    table_top = meta_top - 46
    day_width = 58
    slot_width = 170
    lunch_width = 34
    header_height = 34
    row_height = 30
    day_x = margin
    time_x = day_x + day_width
    lunch_x = time_x + (slot_width * 2)

    commands.append(build_pdf_rect(day_x, table_top - header_height, day_width, header_height, fill_gray=0.84))
    commands.append(build_pdf_centered_text(day_x, table_top - 20, day_width, 'DAY', font='F2', size=9))

    for slot_index in range(2):
        x = time_x + (slot_index * slot_width)
        commands.append(build_pdf_rect(x, table_top - header_height, slot_width, header_height, fill_gray=0.84))
        commands.append(build_pdf_centered_text(x, table_top - 18, slot_width, TIMETABLE_SLOTS[slot_index], font='F2', size=8))

    commands.append(build_pdf_rect(lunch_x, table_top - header_height - (len(TIMETABLE_DAYS) * row_height), lunch_width, header_height + (len(TIMETABLE_DAYS) * row_height), fill_gray=0.88))
    lunch_letters = list('LUNCH')
    for index, letter in enumerate(lunch_letters):
        y = table_top - 18 - (index * 14)
        commands.append(build_pdf_centered_text(lunch_x, y, lunch_width, letter, font='F2', size=8))

    after_lunch_x = lunch_x + lunch_width
    for slot_index in range(2, 4):
        x = after_lunch_x + ((slot_index - 2) * slot_width)
        commands.append(build_pdf_rect(x, table_top - header_height, slot_width, header_height, fill_gray=0.84))
        commands.append(build_pdf_centered_text(x, table_top - 18, slot_width, TIMETABLE_SLOTS[slot_index], font='F2', size=8))

    for day_order, day_name in enumerate(TIMETABLE_DAYS):
        row_top = table_top - header_height - (day_order * row_height)
        row_bottom = row_top - row_height
        commands.append(build_pdf_rect(day_x, row_bottom, day_width, row_height))
        commands.append(build_pdf_centered_text(day_x, row_bottom + 11, day_width, day_name[:3].upper(), font='F2', size=9))

        for slot_index in range(2):
            x = time_x + (slot_index * slot_width)
            commands.append(build_pdf_rect(x, row_bottom, slot_width, row_height))
            entry = entry_map.get((day_order, slot_index))
            if entry:
                commands.append(build_pdf_centered_text(x, row_bottom + 16, slot_width, truncate_text(entry.subject_name or entry.subject_code, 20), font='F2', size=8))
                commands.append(build_pdf_centered_text(x, row_bottom + 6, slot_width, truncate_text(entry.faculty_name or '-', 16), size=7))
            else:
                commands.append(build_pdf_centered_text(x, row_bottom + 11, slot_width, truncate_text(get_non_teaching_label(day_order, slot_index), 16), size=7))

        for slot_index in range(2, 4):
            x = after_lunch_x + ((slot_index - 2) * slot_width)
            commands.append(build_pdf_rect(x, row_bottom, slot_width, row_height))
            entry = entry_map.get((day_order, slot_index))
            if entry:
                commands.append(build_pdf_centered_text(x, row_bottom + 16, slot_width, truncate_text(entry.subject_name or entry.subject_code, 20), font='F2', size=8))
                commands.append(build_pdf_centered_text(x, row_bottom + 6, slot_width, truncate_text(entry.faculty_name or '-', 16), size=7))
            else:
                commands.append(build_pdf_centered_text(x, row_bottom + 11, slot_width, truncate_text(get_non_teaching_label(day_order, slot_index), 16), size=7))

    subjects_top = table_top - header_height - (len(TIMETABLE_DAYS) * row_height) - 18
    code_w = 90
    subject_w = 420
    faculty_w = content_width - code_w - subject_w
    subject_header_h = 24
    subject_row_h = 20
    commands.append(build_pdf_rect(margin, subjects_top - subject_header_h, code_w, subject_header_h, fill_gray=0.84))
    commands.append(build_pdf_rect(margin + code_w, subjects_top - subject_header_h, subject_w, subject_header_h, fill_gray=0.84))
    commands.append(build_pdf_rect(margin + code_w + subject_w, subjects_top - subject_header_h, faculty_w, subject_header_h, fill_gray=0.84))
    commands.append(build_pdf_centered_text(margin, subjects_top - 16, code_w, 'SUBJECT CODE', font='F2', size=9))
    commands.append(build_pdf_centered_text(margin + code_w, subjects_top - 16, subject_w, 'SUBJECT NAME', font='F2', size=9))
    commands.append(build_pdf_centered_text(margin + code_w + subject_w, subjects_top - 16, faculty_w, 'FACULTY NAME', font='F2', size=9))

    for row_index, (subject_code, subject_name, faculty_name) in enumerate(unique_subjects[:8]):
        row_top = subjects_top - subject_header_h - (row_index * subject_row_h)
        row_bottom = row_top - subject_row_h
        commands.append(build_pdf_rect(margin, row_bottom, code_w, subject_row_h))
        commands.append(build_pdf_rect(margin + code_w, row_bottom, subject_w, subject_row_h))
        commands.append(build_pdf_rect(margin + code_w + subject_w, row_bottom, faculty_w, subject_row_h))
        commands.append(build_pdf_text(margin + 6, row_bottom + 7, truncate_text(subject_code, 14), font='F2', size=8))
        commands.append(build_pdf_text(margin + code_w + 6, row_bottom + 7, truncate_text(subject_name, 52), size=8))
        commands.append(build_pdf_text(margin + code_w + subject_w + 6, row_bottom + 7, truncate_text(faculty_name, 24), size=8))

    footer_y = margin + 10
    commands.append(build_pdf_text(margin, footer_y, 'Generated from Attendance Management System', size=8))

    return build_pdf_document('\n'.join(commands), page_width=page_width, page_height=page_height)


def get_section_subject_format(batch_id, semester_no):
    return SectionSubjectFormat.query.filter_by(batch_id=batch_id, semester=semester_no).first()


def get_section_subject_assignments(section):
    return SectionSubjectAssignment.query.filter_by(section_id=section.section_id, semester=section.current_semester).all()


def get_section_semester_subjects(dept_id, section):
    assigned = SectionSubjectAssignment.query.filter_by(
        section_id=section.section_id,
        semester=section.current_semester
    ).all()
    if not assigned:
        return []
    assigned_codes = [a.subject_code for a in assigned]
    assigned_subjects = Subject.query.filter(
        Subject.subject_code.in_(assigned_codes),
        Subject.dept_id == dept_id
    ).all()
    by_code = {subject.subject_code: subject for subject in assigned_subjects}
    ordered = []
    seen = set()
    for item in assigned:
        subject = by_code.get(item.subject_code)
        if not subject or subject.subject_code in seen:
            continue
        ordered.append(subject)
        seen.add(subject.subject_code)
    return ordered


def get_section_subject_allocations(section):
    allocations = (
        FacultyBatchSection.query
        .filter_by(section_id=section.section_id, batch_id=section.batch_id)
        .all()
    )
    return {allocation.subject_code: allocation for allocation in allocations}


def get_weekly_class_target(subject_count):
    if subject_count <= 0:
        return 0
    return max(3, min(4, (len(TIMETABLE_DAYS) * len(TIMETABLE_SLOTS)) // subject_count))


def validate_section_timetable_readiness(section, dept_id):
    semester_subjects = get_section_semester_subjects(dept_id, section)
    if not semester_subjects:
        return {
            'ready': False,
            'error': 'No subjects assigned for this section. Please assign subjects first.',
            'subjects': [],
            'allocations': {},
            'missing_subjects': [],
            'weekly_class_target': 0,
        }
    allocation_map = get_section_subject_allocations(section)
    missing_subjects = [subject.subject_code for subject in semester_subjects if subject.subject_code not in allocation_map]
    return {
        'ready': len(missing_subjects) == 0,
        'error': None if not missing_subjects else 'Allocate faculty for all semester subjects before generating the timetable',
        'subjects': semester_subjects,
        'allocations': allocation_map,
        'missing_subjects': missing_subjects,
        'weekly_class_target': get_weekly_class_target(len(semester_subjects)),
    }


def get_busy_faculty_slots(exclude_section_id=None):
    busy_slots = defaultdict(set)
    query = TimetableEntry.query
    if exclude_section_id is not None:
        query = query.filter(TimetableEntry.section_id != exclude_section_id)
    for entry in query.all():
        busy_slots[(entry.day_order, entry.slot_index)].add(entry.faculty_id)
    return busy_slots


def get_non_teaching_label(day_order, slot_index):
    return NON_TEACHING_LABELS[(day_order * len(TIMETABLE_SLOTS) + slot_index) % len(NON_TEACHING_LABELS)]


def build_timetable_grid(entries):
    entry_map = {(entry.day_order, entry.slot_index): entry for entry in entries}
    grid = []
    for day_order, day_name in enumerate(TIMETABLE_DAYS):
        row = {'day_order': day_order, 'day_name': day_name, 'slots': []}
        for slot_index, slot_label in enumerate(TIMETABLE_SLOTS):
            entry = entry_map.get((day_order, slot_index))
            row['slots'].append({
                'day_order': day_order,
                'day_name': day_name,
                'slot_index': slot_index,
                'slot_label': slot_label,
                'subject_code': entry.subject_code if entry else None,
                'subject_name': getattr(entry, 'subject_name', None) if entry else None,
                'faculty_id': entry.faculty_id if entry else None,
                'faculty_name': getattr(entry, 'faculty_name', None) if entry else None,
                'activity_label': None if entry else get_non_teaching_label(day_order, slot_index),
            })
        grid.append(row)
    return grid


def enrich_timetable_entries(entries):
    enriched = []
    for entry in entries:
        subject = db.session.get(Subject, entry.subject_code)
        faculty = db.session.get(Faculty, entry.faculty_id)
        user = db.session.get(User, faculty.user_id) if faculty else None
        entry.subject_name = subject.subject_name if subject else None
        entry.faculty_name = user.name if user else None
        enriched.append(entry)
    return enriched


def _is_lab_subject(subject_type, subject_name=''):
    return 'lab' in (subject_type or '').lower() or 'lab' in (subject_name or '').lower()


def _has_consecutive_theory_conflict(assigned_slots, faculty_id, day_order, slot_index):
    for adj in [slot_index - 1, slot_index + 1]:
        if 0 <= adj < len(TIMETABLE_SLOTS):
            neighbour = assigned_slots.get((day_order, adj))
            if neighbour and neighbour['faculty_id'] == faculty_id and not neighbour.get('is_lab'):
                return True
    return False


def generate_section_timetable(section, subjects, allocation_map):
    import random
    from collections import Counter

    session_templates = []

    for subject in subjects:
        allocation = allocation_map[subject.subject_code]
        is_lab = _is_lab_subject(subject.subject_type, subject.subject_name)
        count = 1 if is_lab else 2
        for _ in range(count):
            session_templates.append({
                'subject_code': subject.subject_code,
                'faculty_id': allocation.faculty_id,
                'is_lab': is_lab,
            })

    random.shuffle(session_templates)

    busy_slots = get_busy_faculty_slots(exclude_section_id=section.section_id)
    assigned_slots = {}

    def has_same_subject_same_day(subject_code, day_order):
        for (d, _), session in assigned_slots.items():
            if d == day_order and session['subject_code'] == subject_code:
                return True
        return False

    def candidate_slots(session):
        slots = []
        for day_order in range(len(TIMETABLE_DAYS)):
            for slot_index in range(len(TIMETABLE_SLOTS)):
                slot_key = (day_order, slot_index)
                if slot_key in assigned_slots:
                    continue
                if session['faculty_id'] in busy_slots.get(slot_key, set()):
                    continue
                if not session['is_lab'] and _has_consecutive_theory_conflict(
                        assigned_slots, session['faculty_id'], day_order, slot_index):
                    continue
                if has_same_subject_same_day(session['subject_code'], day_order):
                    continue
                slots.append(slot_key)
        random.shuffle(slots)
        return slots

    def backtrack(index):
        if index == len(session_templates):
            return True
        session = session_templates[index]
        for day_order, slot_index in candidate_slots(session):
            assigned_slots[(day_order, slot_index)] = session
            if backtrack(index + 1):
                return True
            del assigned_slots[(day_order, slot_index)]
        return False

    if not backtrack(0):
        raise ValueError("Unable to generate timetable with given constraints")

    generated_entries = []
    for (day_order, slot_index), session in assigned_slots.items():
        generated_entries.append(
            TimetableEntry(
                section_id=section.section_id,
                batch_id=section.batch_id,
                faculty_id=session['faculty_id'],
                subject_code=session['subject_code'],
                day_order=day_order,
                day_name=TIMETABLE_DAYS[day_order],
                slot_index=slot_index,
                slot_label=TIMETABLE_SLOTS[slot_index],
            )
        )

    return generated_entries, None


def save_section_timetable(section, subjects, allocation_map, raw_entries):
    weekly_target = get_weekly_class_target(len(subjects))
    subject_codes = {subject.subject_code for subject in subjects}
    subject_type_map = {subject.subject_code: subject.subject_type for subject in subjects}
    subject_name_map = {subject.subject_code: subject.subject_name for subject in subjects}
    busy_slots = get_busy_faculty_slots(exclude_section_id=section.section_id)
    seen_slots = set()
    subject_counts = Counter()
    prepared_entries = []
    slot_faculty_map = {}

    for item in raw_entries:
        subject_code = (item.get('subject_code') or '').strip().upper()
        if not subject_code:
            continue
        day_order = item.get('day_order')
        slot_index = item.get('slot_index')
        if day_order is None or slot_index is None:
            raise ValueError('Each timetable entry must include day_order and slot_index')
        try:
            day_order = int(day_order)
            slot_index = int(slot_index)
        except (TypeError, ValueError):
            raise ValueError('Invalid timetable slot coordinates')

        if day_order < 0 or day_order >= len(TIMETABLE_DAYS) or \
           slot_index < 0 or slot_index >= len(TIMETABLE_SLOTS):
            raise ValueError('Timetable slot is out of range')

        if subject_code not in subject_codes:
            raise ValueError(f'{subject_code} is not a valid subject for this section')
        if subject_code not in allocation_map:
            raise ValueError(f'{subject_code} is not allocated to any faculty for this section')

        slot_key = (day_order, slot_index)
        faculty_id = allocation_map[subject_code].faculty_id

        if slot_key in seen_slots:
            raise ValueError('A timetable slot cannot contain more than one subject')

        if faculty_id in busy_slots.get(slot_key, set()):
            raise ValueError(
                f'Faculty conflict: the faculty teaching {subject_code} is already '
                f'scheduled in another section on {TIMETABLE_DAYS[day_order]} '
                f'{TIMETABLE_SLOTS[slot_index]}'
            )

        is_lab = _is_lab_subject(
            subject_type_map.get(subject_code, ''),
            subject_name_map.get(subject_code, ''),
        )
        if not is_lab and _has_consecutive_theory_conflict(slot_faculty_map, faculty_id, day_order, slot_index):
            raise ValueError(
                f'Consecutive theory conflict: {subject_code} on '
                f'{TIMETABLE_DAYS[day_order]} slot {slot_index + 1} would give '
                f'the same faculty two back-to-back theory classes.'
            )

        seen_slots.add(slot_key)
        slot_faculty_map[slot_key] = {'faculty_id': faculty_id, 'is_lab': is_lab}
        subject_counts[subject_code] += 1

        prepared_entries.append(
            TimetableEntry(
                section_id=section.section_id,
                batch_id=section.batch_id,
                faculty_id=faculty_id,
                subject_code=subject_code,
                day_order=day_order,
                day_name=TIMETABLE_DAYS[day_order],
                slot_index=slot_index,
                slot_label=TIMETABLE_SLOTS[slot_index],
            )
        )

    expected_counts = {}
    for subject in subjects:
        is_lab = _is_lab_subject(subject.subject_type, subject.subject_name)
        expected_counts[subject.subject_code] = 1 if is_lab else 2
    if dict(subject_counts) != expected_counts:
        raise ValueError(
            f'Each subject must appear exactly {weekly_target} times in the timetable'
        )

    TimetableEntry.query.filter_by(section_id=section.section_id).delete()
    db.session.add_all(prepared_entries)
    return weekly_target


def sync_batch_active_semester(batch_id, semester_no):
    Semester.query.filter_by(batch_id=batch_id).update({'is_active': False})
    active_semester = Semester.query.filter_by(batch_id=batch_id, semester_no=semester_no).first()
    if not active_semester:
        active_semester = Semester(batch_id=batch_id, semester_no=semester_no, is_active=True)
        db.session.add(active_semester)
    else:
        active_semester.is_active = True
    return active_semester

@dept_bp.route('/batches', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_batches(current_user):
    batches = (
        Batch.query
        .filter_by(dept_id=current_user.dept_id)
        .order_by(Batch.batch_name.desc(), Batch.batch_id.asc())
        .all()
    )
    result = []
    for batch in batches:
        program = db.session.get(Program, batch.program_id)
        active_semester = Semester.query.filter_by(batch_id=batch.batch_id, is_active=True).first()
        result.append({
            'id': batch.batch_id,
            'name': batch.batch_name,
            'program_id': batch.program_id,
            'program_name': program.program_name if program else None,
            'current_semester': active_semester.semester_no if active_semester else None,
        })
    return jsonify(result)

@dept_bp.route('/batches', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def create_batch(current_user):
    data = request.get_json()
    required = ['name', 'program_id']
    if not data or not all(data.get(k) for k in required):
        return jsonify({'error': 'name and program_id are required'}), 400
    existing = Batch.query.filter_by(batch_name=data['name'], dept_id=current_user.dept_id).first()
    if existing:
        return jsonify({'error': 'Batch name already exists in this department'}), 409

    program = db.session.get(Program, data['program_id'])
    if not program:
        return jsonify({'error': 'Invalid program_id'}), 400

    # Academic range validation: e.g., 2023-25 => 2-year, 2021-25 => 4-year
    year_match = re.match(r'^(?P<start>\d{4})-(?P<end>\d{2,4})$', str(data['name']).strip())
    if year_match:
        start_year = int(year_match.group('start'))
        end_part = year_match.group('end')
        end_year = int(end_part) if len(end_part) == 4 else int(str(start_year)[:2] + end_part)
        year_span = end_year - start_year
        expected_year_span = program.duration_semesters // 2
        if year_span != expected_year_span:
            return jsonify({'error': f'Batch duration mismatch: program is {program.duration_semesters} semesters, so batch, e.g. should span {expected_year_span} years.'}), 400

    batch = Batch(batch_name=data['name'], dept_id=current_user.dept_id, program_id=data['program_id'])
    db.session.add(batch)
    db.session.flush()
    duration = program.duration_semesters
    for semester_no in range(1, duration + 1):
        db.session.add(Semester(batch_id=batch.batch_id, semester_no=semester_no, is_active=(semester_no == 1)))
    db.session.commit()
    return jsonify({'id': batch.batch_id, 'name': batch.batch_name}), 201


@dept_bp.route('/sections/delete_group', methods=['DELETE'])
@token_required
@role_required('DEPT_ADMIN')
def delete_group_sections(current_user):
    data = request.get_json() or {}
    batch_id = data.get('batch_id')
    program_id = data.get('program_id')
    if not batch_id or not program_id:
        return jsonify({'error': 'batch_id and program_id are required'}), 400
    sections = Section.query.filter_by(batch_id=batch_id, program_id=program_id, dept_id=current_user.dept_id).all()
    if not sections:
        return jsonify({'message': 'No sections found for this batch/program', 'deleted': 0})
    deleted = len(sections)
    for s in sections:
        # also clear related students and timetable entries in this section
        Student.query.filter_by(section_id=s.section_id).delete()
        TimetableEntry.query.filter_by(section_id=s.section_id).delete()
        db.session.delete(s)
    db.session.commit()
    return jsonify({'message': f'Deleted {deleted} section(s) for batch/program', 'deleted': deleted})

@dept_bp.route('/programs', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_programs(current_user):
    return jsonify([{'id': p.program_id, 'name': p.program_name, 'duration_semesters': p.duration_semesters} for p in Program.query.all()])

@dept_bp.route('/programs', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def create_program(current_user):
    data = request.get_json()
    if not data or not data.get('name'):
        return jsonify({'error': 'Program name required'}), 400
    program = Program(program_name=data['name'], duration_semesters=data.get('duration_semesters', 8))
    db.session.add(program)
    db.session.commit()
    return jsonify({'id': program.program_id, 'name': program.program_name, 'duration_semesters': program.duration_semesters}), 201

@dept_bp.route('/stats', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_stats(current_user):
    section_count = Section.query.filter_by(dept_id=current_user.dept_id).count()
    subject_count = Subject.query.filter_by(dept_id=current_user.dept_id).count()
    faculty_count = Faculty.query.filter_by(dept_id=current_user.dept_id).count()
    student_count = Student.query.filter_by(dept_id=current_user.dept_id).count()
    allocation_count = FacultyBatchSection.query.join(Faculty, FacultyBatchSection.faculty_id == Faculty.faculty_id).filter(Faculty.dept_id == current_user.dept_id).count()
    return jsonify({
        'sections': section_count,
        'subjects': subject_count,
        'faculty': faculty_count,
        'students': student_count,
        'allocations': allocation_count,
    })

# ── Sections CRUD ─────────────────────────────────────
@dept_bp.route('/sections', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_sections(current_user):
    sections = (
        Section.query
        .join(Batch, Section.batch_id == Batch.batch_id)
        .filter_by(dept_id=current_user.dept_id)
        .order_by(Batch.batch_name.asc(), Section.current_semester.asc(), Section.section_name.asc())
        .all()
    )

    result = []
    now = datetime.now(timezone.utc)
    for s in sections:
        batch = db.session.get(Batch, s.batch_id)
        program = db.session.get(Program, s.program_id or (batch.program_id if batch else None))
        recently_updated = False
        if s.updated_at:
            try:
                recently_updated = (now - s.updated_at).days <= 30
            except Exception:
                recently_updated = False
        result.append({
            'id': s.section_id,
            'name': s.section_name,
            'batch_id': s.batch_id,
            'batch_name': batch.batch_name if batch else None,
            'program_id': program.program_id if program else None,
            'program_name': program.program_name if program else None,
            'current_semester': s.current_semester,
            'updated_at': s.updated_at.isoformat() if s.updated_at else None,
            'recently_updated': recently_updated,
        })
    return jsonify(result)

@dept_bp.route('/sections', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def create_section(current_user):
    data = request.get_json()
    if not data or not data.get('name') or not data.get('batch_id'):
        return jsonify({'error': 'Name and batch_id are required'}), 400
    # Check if section name already exists in this batch
    existing = Section.query.filter_by(section_name=data['name'], batch_id=data['batch_id']).first()
    if existing:
        return jsonify({'error': 'Section already exists in this batch'}), 409
    batch = db.session.get(Batch, data['batch_id'])
    if not batch or batch.dept_id != current_user.dept_id:
        return jsonify({'error': 'Invalid batch'}), 400
    program = db.session.get(Program, batch.program_id)
    if not program:
        return jsonify({'error': 'Invalid program associated with selected batch'}), 400

    existing_batch_section = Section.query.filter_by(batch_id=batch.batch_id, dept_id=current_user.dept_id).first()
    section_semester = existing_batch_section.current_semester if existing_batch_section else data.get('current_semester', 1)

    section = Section(
        section_name=data['name'],
        current_semester=section_semester,
        batch_id=data['batch_id'],
        dept_id=current_user.dept_id,
        program_id=batch.program_id,
    )
    db.session.add(section)
    sync_batch_active_semester(section.batch_id, section.current_semester or 1)
    db.session.commit()
    result = section.to_dict()
    result['batch_name'] = batch.batch_name if batch else None
    result['program_name'] = program.program_name if program else None
    return jsonify(result), 201

@dept_bp.route('/sections/<int:section_id>', methods=['PUT'])
@token_required
@role_required('DEPT_ADMIN')
def update_section(current_user, section_id):
    section = db.session.get(Section, section_id)
    if not section or section.dept_id != current_user.dept_id:
        return jsonify({'error': 'Section not found'}), 404
    data = request.get_json() or {}
    old_semester = section.current_semester
    old_batch_id = section.batch_id

    if data.get('increment_semester'):
        section.current_semester = (section.current_semester or 1) + 1
    elif data.get('current_semester') is not None:
        section.current_semester = int(data['current_semester'])

    if data.get('batch_id') is not None:
        batch = db.session.get(Batch, data['batch_id'])
        if not batch or batch.dept_id != current_user.dept_id:
            return jsonify({'error': 'Invalid batch'}), 400
        section.batch_id = data['batch_id']

    if data.get('program_id') is not None:
        program = db.session.get(Program, data['program_id'])
        if not program:
            return jsonify({'error': 'Invalid program'}), 400
        section.program_id = data['program_id']

    semester_changed = old_semester != section.current_semester
    batch_changed = old_batch_id != section.batch_id
    if semester_changed or batch_changed:
        clear_batch_subject_setup(old_batch_id, section_ids=[section.section_id])
        if batch_changed:
            clear_batch_subject_setup(section.batch_id, section_ids=[section.section_id])

    section.updated_at = datetime.now(timezone.utc)
    TimetableEntry.query.filter_by(section_id=section.section_id).delete()
    sync_batch_active_semester(section.batch_id, section.current_semester or 1)
    db.session.commit()
    batch = db.session.get(Batch, section.batch_id)
    program = db.session.get(Program, section.program_id or (batch.program_id if batch else None))

    result = section.to_dict()
    result['batch_name'] = batch.batch_name if batch else None
    result['program_name'] = program.program_name if program else None
    return jsonify(result)

@dept_bp.route('/sections/<int:section_id>', methods=['DELETE'])
@token_required
@role_required('DEPT_ADMIN')
def delete_section(current_user, section_id):
    section = db.session.get(Section, section_id)
    if not section or section.dept_id != current_user.dept_id:
        return jsonify({'error': 'Section not found'}), 404
    db.session.delete(section)
    db.session.commit()
    return jsonify({'message': 'Section deleted'})

@dept_bp.route('/sections/flat', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_sections_flat(current_user):
    # Return flat list of sections for components that need individual sections
    sections = Section.query.filter_by(dept_id=current_user.dept_id).all()
    result = []
    for section in sections:
        batch = db.session.get(Batch, section.batch_id)
        result.append({
            'id': section.section_id,
            'name': section.section_name,
            'batch_id': section.batch_id,
            'batch_name': batch.batch_name if batch else None,
            'current_semester': section.current_semester,
            'dept_id': section.dept_id,
        })
    return jsonify(result)

@dept_bp.route('/batches/<int:batch_id>/sections', methods=['PUT'])
@token_required
@role_required('DEPT_ADMIN')
def update_batch_sections(current_user, batch_id):
    """Update semester for all sections in a batch"""
    batch = db.session.get(Batch, batch_id)
    if not batch or batch.dept_id != current_user.dept_id:
        return jsonify({'error': 'Batch not found'}), 404
    data = request.get_json()
    if not data or not data.get('current_semester'):
        return jsonify({'error': 'current_semester is required'}), 400
    new_semester = int(data['current_semester'])
    sections = Section.query.filter_by(batch_id=batch_id, dept_id=current_user.dept_id).all()
    semester_changed = any((section.current_semester or 1) != new_semester for section in sections)
    if semester_changed:
        section_ids = [section.section_id for section in sections]
        clear_batch_subject_setup(batch_id, section_ids=section_ids)

    for s in sections:
        s.current_semester = new_semester
        TimetableEntry.query.filter_by(section_id=s.section_id).delete()
    sync_batch_active_semester(batch_id, new_semester)
    db.session.commit()
    return jsonify({'message': f'Updated {len(sections)} sections', 'count': len(sections)})

@dept_bp.route('/batches/<int:batch_id>/sections', methods=['DELETE'])
@token_required
@role_required('DEPT_ADMIN')
def delete_batch_sections(current_user, batch_id):
    """Delete all sections in a batch"""
    batch = db.session.get(Batch, batch_id)
    if not batch or batch.dept_id != current_user.dept_id:
        return jsonify({'error': 'Batch not found'}), 404
    sections = Section.query.filter_by(batch_id=batch_id, dept_id=current_user.dept_id).all()
    count = len(sections)
    for s in sections:
        db.session.delete(s)
    db.session.commit()
    return jsonify({'message': f'Deleted {count} sections', 'count': count})

@dept_bp.route('/subjects', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_subjects(current_user):
    query = Subject.query.filter_by(dept_id=current_user.dept_id)
    batch_id = request.args.get('batch_id')
    program_id = request.args.get('program_id')
    section_id = request.args.get('section_id')
    if batch_id:
        try:
            query = query.filter_by(batch_id=int(batch_id))
        except ValueError:
            pass
    if program_id:
        try:
            query = query.filter_by(program_id=int(program_id))
        except ValueError:
            pass
    subject_type = request.args.get('subject_type') or request.args.get('type')
    if subject_type:
        query = query.filter_by(subject_type=subject_type)
    if section_id:
        try:
            selected_section = db.session.get(Section, int(section_id))
            if not selected_section or selected_section.dept_id != current_user.dept_id:
                return jsonify({'error': 'Section not found'}), 404
            assigned_codes = db.session.query(SectionSubjectAssignment.subject_code).filter_by(
                section_id=selected_section.section_id,
                semester=selected_section.current_semester
            ).distinct().subquery()
            query = query.filter(Subject.subject_code.in_(assigned_codes))
        except ValueError:
            pass

    subjects = query.order_by(Subject.subject_type.asc(), Subject.subject_code.asc()).all()
    return jsonify([{
        'code': s.subject_code,
        'name': s.subject_name,
        'semester': s.semester,
        'credits': s.credits,
        'periods': s.periods,
        'subject_type': s.subject_type,
        'batch_id': s.batch_id,
        'program_id': s.program_id,
    } for s in subjects])


@dept_bp.route('/subjects/types', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_subject_types(current_user):
    types = {normalize_subject_type(row[0]) for row in db.session.query(Subject.subject_type).filter_by(dept_id=current_user.dept_id).distinct().all() if row[0]}
    for default_type in DEFAULT_SUBJECT_TYPES:
        types.add(default_type)
    return jsonify(sorted(types))


@dept_bp.route('/subjects', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def create_subject(current_user):
    data = request.get_json()
    required = ['name', 'subject_type', 'credits', 'periods']
    if not data or not all(data.get(k) for k in required):
        return jsonify({'error': 'name, subject_type, credits, periods are required'}), 400
    normalized_type = normalize_subject_type(data.get('subject_type'))
    if not normalized_type:
        return jsonify({'error': 'subject_type is required'}), 400

    dept_batch_ids = [row[0] for row in db.session.query(Batch.batch_id).filter_by(dept_id=current_user.dept_id).all()]
    dept_active = Semester.query.filter(
        Semester.batch_id.in_(dept_batch_ids),
        Semester.is_active.is_(True)
    ).order_by(Semester.batch_id.asc()).first() if dept_batch_ids else None
    semester = dept_active.semester_no if dept_active else 1

    provided_code = (data.get('code') or '').strip().upper()
    if provided_code:
        code = provided_code
        if db.session.get(Subject, code):
            return jsonify({'error': 'Subject code already exists'}), 409
    else:
        department = db.session.get(Department, current_user.dept_id)
        code = generate_subject_code_for_dept(department.dept_name if department else '', semester)

    subject = Subject(
        subject_code=code,
        subject_name=data['name'],
        semester=None,
        credits=float(data['credits']),
        periods=int(data['periods']),
        subject_type=normalized_type,
        dept_id=current_user.dept_id,
        batch_id=None,
        program_id=None,
    )
    db.session.add(subject)
    db.session.commit()
    return jsonify({'subject_code': subject.subject_code, 'subject_name': subject.subject_name, 'semester': subject.semester, 'credits': subject.credits, 'subject_type': subject.subject_type}), 201


def clean_subject_cell(value, default=''):
    if value is None:
        return default
    return str(value).strip()


def normalize_bulk_header(value):
    return re.sub(r'[^a-z0-9]+', '_', clean_subject_cell(value).lower()).strip('_')


def get_bulk_value(row, *keys):
    normalized = {normalize_bulk_header(key): value for key, value in row.items()}
    for key in keys:
        normalized_key = normalize_bulk_header(key)
        if normalized_key in normalized:
            return normalized[normalized_key]
    return None


def apply_subject_values(subject, data, current_user):
    normalized_type = normalize_subject_type(data.get('subject_type'))
    if not normalized_type:
        raise ValueError('subject_type is required')
    subject.subject_name = clean_subject_cell(data.get('name'))
    subject.subject_type = normalized_type
    subject.credits = float(data.get('credits') or 0)
    subject.periods = int(data.get('periods') or 1)
    subject.batch_id = None
    subject.program_id = None


@dept_bp.route('/subjects/<subject_code>', methods=['PUT'])
@token_required
@role_required('DEPT_ADMIN')
def update_subject(current_user, subject_code):
    subject = db.session.get(Subject, subject_code.upper())
    if not subject or subject.dept_id != current_user.dept_id:
        return jsonify({'error': 'Subject not found'}), 404
    data = request.get_json() or {}
    if not data.get('name') or not data.get('subject_type'):
        return jsonify({'error': 'name and subject_type are required'}), 400
    try:
        apply_subject_values(subject, data, current_user)
        db.session.commit()
    except ValueError as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 400
    return jsonify({'message': 'Subject updated'})


@dept_bp.route('/subjects/bulk', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def bulk_upload_subjects(current_user):
    upload_file = request.files.get('file')
    if not upload_file:
        return jsonify({'error': 'file is required'}), 400

    filename = upload_file.filename.lower()
    if not filename.endswith(('.csv', '.xlsx', '.xls')):
        return jsonify({'error': 'Only CSV/XLSX files are supported'}), 400

    rows = []
    if filename.endswith('.csv'):
        import csv
        try:
            decoded = upload_file.stream.read().decode('utf-8', errors='replace')
            rows.extend(csv.DictReader(decoded.splitlines()))
        except Exception as exc:
            return jsonify({'error': f'CSV parse error: {exc}'}), 400
    else:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return jsonify({'error': 'openpyxl is required for Excel upload'}), 500
        try:
            wb = load_workbook(upload_file, data_only=True)
            sheet = wb.active
            headers = [normalize_bulk_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: row[i] for i in range(len(headers)) if headers[i]})
        except Exception as exc:
            return jsonify({'error': f'Excel parse error: {exc}'}), 400

    created = 0
    updated = 0
    errors = []
    for idx, row in enumerate(rows, start=1):
        try:
            if not any(clean_subject_cell(value) for value in row.values()):
                continue
            data = {
                'code': clean_subject_cell(get_bulk_value(row, 'code', 'subject_code', 'subject code')).upper(),
                'name': clean_subject_cell(get_bulk_value(row, 'name', 'subject_name', 'subject name')),
                'subject_type': clean_subject_cell(get_bulk_value(row, 'subject_type', 'type', 'subject type'), 'Common'),
                'credits': get_bulk_value(row, 'credits', 'credit') or 1,
                'periods': get_bulk_value(row, 'periods', 'period', 'classes', 'no of classes') or 1,
            }
            if not data['name']:
                errors.append({'row': idx, 'error': 'Missing subject name'})
                continue

            subject = db.session.get(Subject, data['code']) if data['code'] else None
            if subject:
                if subject.dept_id != current_user.dept_id:
                    errors.append({'row': idx, 'error': 'Subject code belongs to another department'})
                    continue
                apply_subject_values(subject, data, current_user)
                updated += 1
                continue

            code = data['code']
            if code and db.session.get(Subject, code):
                errors.append({'row': idx, 'error': 'Subject code already exists'})
                continue
            if not code:
                semester = 1
                dept_batch_ids = [row[0] for row in db.session.query(Batch.batch_id).filter_by(dept_id=current_user.dept_id).all()]
                active_sem = Semester.query.filter(
                    Semester.batch_id.in_(dept_batch_ids),
                    Semester.is_active.is_(True)
                ).order_by(Semester.batch_id.asc()).first() if dept_batch_ids else None
                semester = active_sem.semester_no if active_sem else 1
                department = db.session.get(Department, current_user.dept_id)
                code = generate_subject_code_for_dept(department.dept_name if department else '', semester)

            subject = Subject(subject_code=code, dept_id=current_user.dept_id)
            apply_subject_values(subject, data, current_user)
            db.session.add(subject)
            created += 1
        except Exception as exc:
            db.session.rollback()
            errors.append({'row': idx, 'error': str(exc)})

    db.session.commit()
    return jsonify({'created': created, 'updated': updated, 'errors': errors})

@dept_bp.route('/subjects/<subject_code>', methods=['DELETE'])
@token_required
@role_required('DEPT_ADMIN')
def delete_subject(current_user, subject_code):
    subject = db.session.get(Subject, subject_code)
    if not subject or subject.dept_id != current_user.dept_id:
        return jsonify({'error': 'Subject not found'}), 404
    db.session.delete(subject)
    db.session.commit()
    return jsonify({'message': 'Subject deleted'})


@dept_bp.route('/batch-programs', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_batch_programs(current_user):
    """Get all batch-program combinations with section and semester info"""
    batches = Batch.query.filter_by(dept_id=current_user.dept_id).all()
    result = []
    for batch in batches:
        program = db.session.get(Program, batch.program_id)
        sections = Section.query.filter_by(batch_id=batch.batch_id).all()
        active_sem = Semester.query.filter_by(batch_id=batch.batch_id, is_active=True).first()
        result.append({
            'batch_id': batch.batch_id,
            'batch_name': batch.batch_name,
            'program_id': program.program_id if program else None,
            'program_name': program.program_name if program else None,
            'current_semester': active_sem.semester_no if active_sem else 1,
            'section_count': len(sections),
            'sections': [{'id': s.section_id, 'name': s.section_name, 'semester': s.current_semester} for s in sections],
        })
    return jsonify(result)


@dept_bp.route('/subjects/formats', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def define_subject_format(current_user):
    """Define slot-based subject format for a batch/semester."""
    data = request.get_json() or {}
    batch_id = data.get('batch_id')
    semester = data.get('semester')
    subjects_list = data.get('subjects', [])  # list of {code, type, subject_code?}

    if not batch_id or not semester:
        return jsonify({'error': 'batch_id and semester are required'}), 400

    batch = db.session.get(Batch, batch_id)
    if not batch or batch.dept_id != current_user.dept_id:
        return jsonify({'error': 'Invalid batch'}), 400

    try:
        semester = int(semester)
    except (ValueError, TypeError):
        return jsonify({'error': 'semester must be an integer'}), 400

    # Get or create format
    fmt = SectionSubjectFormat.query.filter_by(batch_id=batch_id, semester=semester).first()
    if not fmt:
        fmt = SectionSubjectFormat(batch_id=batch_id, semester=semester)
        db.session.add(fmt)
        db.session.flush()

    # Clear and rebuild subject list
    FormatSubject.query.filter_by(format_id=fmt.id).delete()

    seen_codes = set()
    for item in subjects_list:
        format_code = item.get('code', '').strip().upper()
        subject_type = normalize_subject_type(item.get('type'))
        mapped_subject_code = (item.get('subject_code') or '').strip().upper() or None

        if not format_code or not subject_type:
            continue
        if format_code in seen_codes:
            continue

        if subject_type == 'Common':
            if not mapped_subject_code:
                return jsonify({'error': f'Common slot {format_code} must include a subject selection'}), 400
            mapped_subject = db.session.get(Subject, mapped_subject_code)
            if not mapped_subject or mapped_subject.dept_id != current_user.dept_id:
                return jsonify({'error': f'Subject {mapped_subject_code} not found'}), 400
            if normalize_subject_type(mapped_subject.subject_type) != 'Common':
                return jsonify({'error': f'Subject {mapped_subject_code} must be Common type'}), 400
        else:
            mapped_subject_code = None

        db.session.add(FormatSubject(
            format_id=fmt.id,
            subject_code=format_code,
            subject_type=subject_type,
            mapped_subject_code=mapped_subject_code,
        ))
        seen_codes.add(format_code)

    # Format updates invalidate section-level optional picks for this batch/semester.
    section_ids = [row[0] for row in db.session.query(Section.section_id).filter_by(batch_id=batch_id).all()]
    if section_ids:
        SectionSubjectAssignment.query.filter(
            SectionSubjectAssignment.section_id.in_(section_ids),
            SectionSubjectAssignment.semester == semester
        ).delete(synchronize_session=False)
        TimetableEntry.query.filter(
            TimetableEntry.section_id.in_(section_ids)
        ).delete(synchronize_session=False)

    fmt.updated_at = datetime.now(timezone.utc)
    db.session.commit()
    return jsonify({'message': 'Format defined', 'format_id': fmt.id}), 201


@dept_bp.route('/subjects/formats/<int:batch_id>/<int:semester>', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_format_details(current_user, batch_id, semester):
    """Get slot format details and available subjects grouped by subject type."""
    batch = db.session.get(Batch, batch_id)
    if not batch or batch.dept_id != current_user.dept_id:
        return jsonify({'error': 'Invalid batch'}), 400

    # Get the format if it exists
    fmt = SectionSubjectFormat.query.filter_by(batch_id=batch_id, semester=semester).first()
    format_subjects = []
    available_types = set()
    if fmt:
        format_items = FormatSubject.query.filter_by(format_id=fmt.id).all()
        for item in format_items:
            mapped_subject = db.session.get(Subject, item.mapped_subject_code) if item.mapped_subject_code else None
            available_types.add(normalize_subject_type(item.subject_type))
            format_subjects.append({
                'code': item.subject_code,
                'type': item.subject_type,
                'subject_code': item.mapped_subject_code,
                'subject_name': mapped_subject.subject_name if mapped_subject else None,
            })
        format_subjects = sorted(format_subjects, key=lambda item: subject_code_sort_key(item.get('code')))

    semester_subjects = Subject.query.filter_by(
        dept_id=current_user.dept_id
    ).order_by(Subject.subject_type.asc(), Subject.subject_code.asc()).all()

    subjects_by_type = defaultdict(list)
    for subject in semester_subjects:
        normalized_type = normalize_subject_type(subject.subject_type)
        available_types.add(normalized_type)
        subjects_by_type[normalized_type].append({
            'code': subject.subject_code,
            'name': subject.subject_name,
            'type': normalized_type,
        })

    return jsonify({
        'format_id': fmt.id if fmt else None,
        'batch_id': batch_id,
        'semester': semester,
        'updated_at': fmt.updated_at.isoformat() if fmt and fmt.updated_at else None,
        'format_subjects': format_subjects,
        'subject_types': sorted(available_types.union(set(DEFAULT_SUBJECT_TYPES))),
        'available_subjects_by_type': dict(subjects_by_type),
    })


@dept_bp.route('/sections/<int:section_id>/assign-subjects', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_section_assignments(current_user, section_id):
    """Get existing assignments for a section"""
    section = db.session.get(Section, section_id)
    if not section or section.dept_id != current_user.dept_id:
        return jsonify({'error': 'Section not found'}), 404

    assignments = SectionSubjectAssignment.query.filter_by(
        section_id=section_id,
        semester=section.current_semester
    ).all()

    return jsonify([{
        'format_code': a.format_code,
        'subject_code': a.subject_code,
        'subject_type': a.subject_type,
    } for a in assignments])


@dept_bp.route('/sections/<int:section_id>/assign-subjects', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def assign_section_subjects(current_user, section_id):
    """Assign actual subjects to format slots for a section."""
    section = db.session.get(Section, section_id)
    if not section or section.dept_id != current_user.dept_id:
        return jsonify({'error': 'Section not found'}), 404

    data = request.get_json() or {}
    mapping_list = data.get('subjects', [])  # list of {format_code, subject_code}

    # Get format to validate
    fmt = SectionSubjectFormat.query.filter_by(
        batch_id=section.batch_id,
        semester=section.current_semester
    ).first()

    if not fmt:
        return jsonify({'error': 'Format is not defined for this batch and semester'}), 400
    format_items = FormatSubject.query.filter_by(format_id=fmt.id).all()
    if not format_items:
        return jsonify({'error': 'Format has no slots defined'}), 400

    mapping_by_slot = {}
    for item in mapping_list:
        slot_code = (item.get('format_code') or item.get('code') or '').strip().upper()
        selected_subject_code = (item.get('subject_code') or '').strip().upper()
        if not slot_code:
            continue
        mapping_by_slot[slot_code] = selected_subject_code

    SectionSubjectAssignment.query.filter_by(
        section_id=section_id,
        semester=section.current_semester
    ).delete(synchronize_session=False)

    inserted = 0
    used_subject_codes = set()
    for format_item in format_items:
        slot_code = format_item.subject_code
        slot_type = normalize_subject_type(format_item.subject_type)
        if slot_type == 'Common':
            selected_subject_code = (format_item.mapped_subject_code or '').strip().upper()
            if not selected_subject_code:
                return jsonify({'error': f'Common slot {slot_code} has no selected subject in format'}), 400
        else:
            selected_subject_code = mapping_by_slot.get(slot_code, '')
            if not selected_subject_code:
                return jsonify({'error': f'Please select a subject for slot {slot_code}'}), 400

        if selected_subject_code in used_subject_codes:
            return jsonify({'error': f'Subject {selected_subject_code} selected more than once'}), 400

        subject = db.session.get(Subject, selected_subject_code)
        if not subject or subject.dept_id != current_user.dept_id:
            return jsonify({'error': f'Subject {selected_subject_code} is invalid'}), 400
        if normalize_subject_type(subject.subject_type) != slot_type:
            return jsonify({'error': f'Subject {selected_subject_code} does not match slot type {slot_type}'}), 400

        db.session.add(SectionSubjectAssignment(
            section_id=section_id,
            semester=section.current_semester,
            subject_code=selected_subject_code,
            format_code=slot_code,
            subject_type=slot_type
        ))
        inserted += 1
        used_subject_codes.add(selected_subject_code)

    # Clear timetable entries as optional subjects changed
    TimetableEntry.query.filter_by(section_id=section_id).delete()
    db.session.commit()

    return jsonify({'message': 'Subjects assigned to section', 'subject_count': inserted}), 201


@dept_bp.route('/subjects/clear', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def clear_subjects(current_user):
    """Delete all subjects, formats, and assignments for the department"""
    assignment_count = SectionSubjectAssignment.query.join(Section).filter(Section.dept_id == current_user.dept_id).delete(synchronize_session=False)
    format_subject_count = FormatSubject.query.join(SectionSubjectFormat, FormatSubject.format_id == SectionSubjectFormat.id).join(Batch, SectionSubjectFormat.batch_id == Batch.batch_id).filter(Batch.dept_id == current_user.dept_id).delete(synchronize_session=False)
    format_count = SectionSubjectFormat.query.join(Batch).filter(Batch.dept_id == current_user.dept_id).delete(synchronize_session=False)
    subject_count = Subject.query.filter_by(dept_id=current_user.dept_id).delete()
    db.session.commit()
    return jsonify({'message': f'Cleared {subject_count} subjects, {format_count} formats, {format_subject_count} format subjects, {assignment_count} assignments', 'subject_count': subject_count, 'format_count': format_count, 'format_subject_count': format_subject_count, 'assignment_count': assignment_count})


@dept_bp.route('/subjects/reset', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def reset_subjects(current_user):
    data = request.get_json() or {}
    subs = data.get('subjects')
    if not isinstance(subs, list):
        return jsonify({'error': 'subjects array required'}), 400
    Subject.query.filter_by(dept_id=current_user.dept_id).delete()
    for item in subs:
        if not all(item.get(k) for k in ['code', 'name', 'credits', 'periods', 'subject_type']):
            continue
        sem = int(item.get('semester')) if item.get('semester') else None
        subject = Subject(
            subject_code=item['code'].strip().upper(),
            subject_name=item['name'].strip(),
            semester=sem,
            credits=float(item['credits']),
            periods=int(item['periods']),
            subject_type=item['subject_type'],
            dept_id=current_user.dept_id,
            batch_id=int(item['batch_id']) if item.get('batch_id') else None,
            program_id=int(item['program_id']) if item.get('program_id') else None,
        )
        db.session.add(subject)
    db.session.commit()
    return jsonify({'message': 'Subject reset completed'})


@dept_bp.route('/students/reset', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def reset_students(current_user):
    data = request.get_json() or {}
    studs = data.get('students')
    if not isinstance(studs, list):
        return jsonify({'error': 'students array required'}), 400
    # delete student-centric tables, then the students
    for st in Student.query.filter_by(dept_id=current_user.dept_id).all():
        AttendanceRecord.query.filter_by(student_id=st.student_id).delete()
        Mark.query.filter_by(student_id=st.student_id).delete()
        db.session.delete(st)
    db.session.commit()
    created = 0
    for item in studs:
        if not all(item.get(k) for k in ['roll_no', 'name', 'batch_id', 'section_id', 'email', 'password']):
            continue
        batch_id = int(item['batch_id'])
        section_id = int(item['section_id'])
        student = Student(
            student_name=item['name'].strip(),
            email=item['email'].strip(),
            roll_no=item['roll_no'].strip(),
            batch_id=batch_id,
            section_id=section_id,
            dept_id=current_user.dept_id,
            student_type=item.get('student_type', 'Regular'),
            passport_number=item.get('passport_number'),
            category=item.get('category', 'General'),
            entrance_marks=float(item.get('entrance_marks') or 0),
        )
        db.session.add(student)
        created += 1
    db.session.commit()
    return jsonify({'message': 'Student reset completed', 'created': created})


# ── Faculty Management ────────────────────────────────
@dept_bp.route('/faculty', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_faculty(current_user):
    faculty_list = []
    for f in Faculty.query.filter_by(dept_id=current_user.dept_id).all():
        user = db.session.get(User, f.user_id)
        allocations = FacultyBatchSection.query.filter_by(faculty_id=f.faculty_id).all()
        # Format allocations as Section-Subject (e.g., A1-CS301)
        formatted_allocs = []
        for a in allocations:
            section = db.session.get(Section, a.section_id)
            batch = db.session.get(Batch, a.batch_id)
            program = db.session.get(Program, batch.program_id) if batch else None
            section_name = section.section_name if section else 'Unknown'
            formatted_allocs.append({
                'batch_id': a.batch_id,
                'section_id': a.section_id,
                'section_name': section_name,
                'subject_code': a.subject_code,
                'display': f'{batch.batch_name if batch else "Unknown"}-{program.program_name if program else "Unknown"}-{section_name}-{a.subject_code}',
            })
        faculty_list.append({
            'id': f.faculty_id,
            'name': user.name if user else None,
            'email': user.email if user else None,
            'phone': user.phone if user else None,
            'designation': user.designation if user else None,
            'dept_id': f.dept_id,
            'allocations': formatted_allocs,
        })
    return jsonify(faculty_list)

@dept_bp.route('/faculty', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def create_faculty(current_user):
    data = request.get_json()
    required = ['name', 'email', 'password']
    if not data or not all(data.get(k) for k in required):
        return jsonify({'error': 'name, email, password are required'}), 400
    email = normalize_email(data['email'])
    if User.query.filter_by(email=email).first():
        return jsonify({'error': 'Email already exists'}), 409
    user = User(name=data['name'], email=email, password_hash=hash_password(data['password']), phone=data.get('phone'), designation=data.get('designation'), role='FACULTY', dept_id=current_user.dept_id, college_id=current_user.college_id)
    db.session.add(user)
    db.session.flush()
    faculty = Faculty(user_id=user.user_id, dept_id=current_user.dept_id)
    db.session.add(faculty)
    db.session.commit()
    return jsonify({'id': faculty.faculty_id, 'name': user.name, 'email': user.email, 'phone': user.phone, 'designation': user.designation, 'dept_id': faculty.dept_id}), 201

@dept_bp.route('/faculty/<int:faculty_id>', methods=['PUT'])
@token_required
@role_required('DEPT_ADMIN')
def update_faculty(current_user, faculty_id):
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or faculty.dept_id != current_user.dept_id:
        return jsonify({'error': 'Faculty not found'}), 404

    user = db.session.get(User, faculty.user_id)
    if not user:
        return jsonify({'error': 'User not found'}), 404

    data = request.get_json() or {}

    if data.get('name'):
        user.name = data['name']
    if data.get('email'):
        email = normalize_email(data['email'])
        if User.query.filter_by(email=email).filter(User.user_id != user.user_id).first():
            return jsonify({'error': 'Email already exists'}), 409
        user.email = email
    if data.get('phone') is not None:
        user.phone = data['phone']
    if data.get('designation') is not None:
        user.designation = data['designation']

    db.session.commit()
    return jsonify({'id': faculty.faculty_id, 'name': user.name, 'email': user.email, 'phone': user.phone, 'designation': user.designation})

@dept_bp.route('/faculty/<int:faculty_id>', methods=['DELETE'])
@token_required
@role_required('DEPT_ADMIN')
def delete_faculty(current_user, faculty_id):
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or faculty.dept_id != current_user.dept_id:
        return jsonify({'error': 'Faculty not found'}), 404
    user = db.session.get(User, faculty.user_id)
    FacultyBatchSection.query.filter_by(faculty_id=faculty_id).delete()
    db.session.delete(faculty)
    if user:
        db.session.delete(user)
    db.session.commit()
    return jsonify({'message': 'Faculty deleted'})

@dept_bp.route('/faculty/<int:faculty_id>/allocations', methods=['DELETE'])
@token_required
@role_required('DEPT_ADMIN')
def clear_faculty_allocations(current_user, faculty_id):
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or faculty.dept_id != current_user.dept_id:
        return jsonify({'error': 'Faculty not found'}), 404
    impacted_section_ids = [alloc.section_id for alloc in FacultyBatchSection.query.filter_by(faculty_id=faculty_id).all()]
    count = FacultyBatchSection.query.filter_by(faculty_id=faculty_id).delete()
    for section_id in impacted_section_ids:
        TimetableEntry.query.filter_by(section_id=section_id).delete()
    db.session.commit()
    return jsonify({'message': f'Cleared {count} allocations'})

@dept_bp.route('/faculty/<int:faculty_id>/timetable', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_faculty_timetable(current_user, faculty_id):
    allocations = FacultyBatchSection.query.filter_by(faculty_id=faculty_id).all()
    output = []
    for a in allocations:
        section = db.session.get(Section, a.section_id)
        batch = db.session.get(Batch, a.batch_id)
        subject = db.session.get(Subject, a.subject_code)
        output.append({
            'batch_id': a.batch_id,
            'batch_name': batch.batch_name if batch else None,
            'section_id': a.section_id,
            'section_name': section.section_name if section else None,
            'subject_code': a.subject_code,
            'subject_name': subject.subject_name if subject else None,
        })
    return jsonify(output)

@dept_bp.route('/faculty/<int:faculty_id>/allocations', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def assign_faculty_allocation(current_user, faculty_id):
    data = request.get_json()
    required = ['batch_id', 'section_id', 'subject_code']
    if not data or not all(data.get(k) for k in required):
        return jsonify({'error': 'batch_id, section_id, subject_code are required'}), 400

    # ensure faculty exists and belongs to department
    faculty = db.session.get(Faculty, faculty_id)
    if not faculty or faculty.dept_id != current_user.dept_id:
        return jsonify({'error': 'Faculty not found'}), 404

    # ensure batch/section belongs to department and section is in batch
    batch = db.session.get(Batch, data['batch_id'])
    section = db.session.get(Section, data['section_id'])
    if (not batch or batch.dept_id != current_user.dept_id or
        not section or section.dept_id != current_user.dept_id or
        section.batch_id != data['batch_id']):
        return jsonify({'error': 'Invalid batch or section'}), 400

    # subject from correct dept and semester
    subject = db.session.get(Subject, data['subject_code'])
    if not subject or subject.dept_id != current_user.dept_id:
        return jsonify({'error': 'Invalid subject'}), 400

    existing_allocation = FacultyBatchSection.query.filter_by(
        batch_id=data['batch_id'],
        section_id=data['section_id'],
        subject_code=data['subject_code'],
    ).first()
    if existing_allocation:
        if existing_allocation.faculty_id == faculty_id:
            return jsonify({'error': 'Allocation exists'}), 409
        return jsonify({'error': 'This subject is already allocated for the section. Delete it before reassigning faculty.'}), 409

    alloc = FacultyBatchSection(faculty_id=faculty_id, batch_id=data['batch_id'], section_id=data['section_id'], subject_code=data['subject_code'])
    db.session.add(alloc)
    TimetableEntry.query.filter_by(section_id=data['section_id']).delete()
    db.session.commit()
    return jsonify({'message': 'Allocation assigned'}), 201

@dept_bp.route('/allocations', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_allocations(current_user):
    allocs = FacultyBatchSection.query.join(Faculty, FacultyBatchSection.faculty_id == Faculty.faculty_id).filter(Faculty.dept_id == current_user.dept_id).all()
    data = []
    for a in allocs:
        faculty = db.session.get(Faculty, a.faculty_id)
        user = db.session.get(User, faculty.user_id) if faculty else None
        section = db.session.get(Section, a.section_id)
        batch = db.session.get(Batch, a.batch_id)
        subject = db.session.get(Subject, a.subject_code)
        data.append({
            'faculty_id': a.faculty_id,
            'faculty_name': user.name if user else None,
            'section_id': a.section_id,
            'section_name': section.section_name if section else None,
            'batch_id': a.batch_id,
            'batch_name': batch.batch_name if batch else None,
            'subject_code': a.subject_code,
            'subject_name': subject.subject_name if subject else None,
        })
    return jsonify(data)

@dept_bp.route('/allocations', methods=['DELETE'])
@token_required
@role_required('DEPT_ADMIN')
def delete_allocation(current_user):
    data = request.get_json()
    required = ['faculty_id', 'batch_id', 'section_id', 'subject_code']
    if not data or not all(data.get(k) for k in required):
        return jsonify({'error': 'faculty_id, batch_id, section_id, subject_code are required'}), 400
    alloc = FacultyBatchSection.query.filter_by(faculty_id=data['faculty_id'], batch_id=data['batch_id'], section_id=data['section_id'], subject_code=data['subject_code']).first()
    if not alloc:
        return jsonify({'error': 'Allocation not found'}), 404
    db.session.delete(alloc)
    TimetableEntry.query.filter_by(section_id=data['section_id']).delete()
    db.session.commit()
    return jsonify({'message': 'Allocation deleted'})


@dept_bp.route('/allocations', methods=['PUT'])
@token_required
@role_required('DEPT_ADMIN')
def reassign_allocation(current_user):
    data = request.get_json()
    required = ['faculty_id', 'batch_id', 'section_id', 'subject_code']
    if not data or not all(data.get(k) for k in required):
        return jsonify({'error': 'faculty_id, batch_id, section_id, subject_code are required'}), 400

    faculty = db.session.get(Faculty, data['faculty_id'])
    if not faculty or faculty.dept_id != current_user.dept_id:
        return jsonify({'error': 'Faculty not found'}), 404

    batch = db.session.get(Batch, data['batch_id'])
    section = db.session.get(Section, data['section_id'])
    if (not batch or batch.dept_id != current_user.dept_id or
        not section or section.dept_id != current_user.dept_id or
        section.batch_id != data['batch_id']):
        return jsonify({'error': 'Invalid batch or section'}), 400

    subject = db.session.get(Subject, data['subject_code'])
    if not subject or subject.dept_id != current_user.dept_id:
        return jsonify({'error': 'Invalid subject'}), 400

    allocation = FacultyBatchSection.query.filter_by(
        batch_id=data['batch_id'],
        section_id=data['section_id'],
        subject_code=data['subject_code'],
    ).first()
    if not allocation:
        return jsonify({'error': 'Allocation not found'}), 404

    if allocation.faculty_id == data['faculty_id']:
        return jsonify({'message': 'Allocation already assigned to this faculty'})

    db.session.delete(allocation)
    db.session.flush()
    db.session.add(FacultyBatchSection(
        faculty_id=data['faculty_id'],
        batch_id=data['batch_id'],
        section_id=data['section_id'],
        subject_code=data['subject_code'],
    ))
    TimetableEntry.query.filter_by(section_id=data['section_id']).delete()
    db.session.commit()
    return jsonify({'message': 'Allocation reassigned successfully'})

@dept_bp.route('/sections/<int:section_id>/timetable', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_section_timetable(current_user, section_id):
    section = db.session.get(Section, section_id)
    if not section or section.dept_id != current_user.dept_id:
        return jsonify({'error': 'Section not found'}), 404

    readiness = validate_section_timetable_readiness(section, current_user.dept_id)
    entries = enrich_timetable_entries(
        TimetableEntry.query
        .filter_by(section_id=section.section_id)
        .order_by(TimetableEntry.day_order.asc(), TimetableEntry.slot_index.asc())
        .all()
    )
    allocation_details = []
    for subject in readiness['subjects']:
        allocation = readiness['allocations'].get(subject.subject_code)
        faculty = db.session.get(Faculty, allocation.faculty_id) if allocation else None
        user = db.session.get(User, faculty.user_id) if faculty else None
        allocation_details.append({
            'subject_code': subject.subject_code,
            'subject_name': subject.subject_name,
            'faculty_id': allocation.faculty_id if allocation else None,
            'faculty_name': user.name if user else None,
        })

    return jsonify({
        'generated': len(entries) > 0,
        'can_generate': readiness['ready'],
        'message': readiness['error'],
        'missing_subjects': readiness['missing_subjects'],
        'weekly_class_target': readiness['weekly_class_target'],
        'allocated_subjects': allocation_details,
        'grid': build_timetable_grid(entries),
    })


@dept_bp.route('/sections/<int:section_id>/timetable/generate', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def generate_timetable(current_user, section_id):
    section = db.session.get(Section, section_id)
    if not section or section.dept_id != current_user.dept_id:
        return jsonify({'error': 'Section not found'}), 404

    readiness = validate_section_timetable_readiness(section, current_user.dept_id)
    if not readiness['ready']:
        return jsonify({'error': readiness['error'], 'missing_subjects': readiness['missing_subjects']}), 400

    TimetableEntry.query.filter_by(section_id=section.section_id).delete()
    try:
        generated_entries, weekly_target = generate_section_timetable(section, readiness['subjects'], readiness['allocations'])
    except ValueError as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 400

    db.session.add_all(generated_entries)
    db.session.commit()
    return jsonify({'message': 'Timetable generated successfully', 'weekly_class_target': weekly_target}), 201


@dept_bp.route('/sections/<int:section_id>/timetable', methods=['PUT'])
@token_required
@role_required('DEPT_ADMIN')
def update_section_timetable(current_user, section_id):
    section = db.session.get(Section, section_id)
    if not section or section.dept_id != current_user.dept_id:
        return jsonify({'error': 'Section not found'}), 404

    readiness = validate_section_timetable_readiness(section, current_user.dept_id)
    if not readiness['ready']:
        return jsonify({'error': readiness['error'], 'missing_subjects': readiness['missing_subjects']}), 400

    data = request.get_json() or {}
    entries = data.get('entries')
    if not isinstance(entries, list):
        return jsonify({'error': 'entries must be provided as a list'}), 400

    try:
        weekly_target = save_section_timetable(section, readiness['subjects'], readiness['allocations'], entries)
        db.session.commit()
    except ValueError as exc:
        db.session.rollback()
        return jsonify({'error': str(exc)}), 400

    return jsonify({'message': 'Timetable updated successfully', 'weekly_class_target': weekly_target})


@dept_bp.route('/sections/<int:section_id>/timetable/download', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def download_section_timetable(current_user, section_id):
    section = db.session.get(Section, section_id)
    if not section or section.dept_id != current_user.dept_id:
        return jsonify({'error': 'Section not found'}), 404

    entries = enrich_timetable_entries(
        TimetableEntry.query
        .filter_by(section_id=section.section_id)
        .order_by(TimetableEntry.day_order.asc(), TimetableEntry.slot_index.asc())
        .all()
    )
    if not entries:
        return jsonify({'error': 'Timetable has not been generated yet'}), 404

    pdf_bytes = build_timetable_pdf(section, entries)
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=section_{section.section_name}_{section.batch_id}_timetable.pdf'
    return response


def calculate_subject_total(marks_by_exam):
    mid1 = marks_by_exam.get('mid1', 0)
    mid2 = marks_by_exam.get('mid2', 0)
    assignment1 = marks_by_exam.get('assignment1', 0)
    assignment2 = marks_by_exam.get('assignment2', 0)
    return round(min(max(mid1, mid2), 20) + min(assignment1 + assignment2, 10), 2)


def get_class_report_payload(current_user, section_id, report_type):
    section = db.session.get(Section, section_id)
    if not section or section.dept_id != current_user.dept_id:
        raise ValueError('Section not found')
    if report_type not in ('attendance', 'marks'):
        raise ValueError('Invalid report type')

    batch = db.session.get(Batch, section.batch_id)
    subjects = get_section_semester_subjects(current_user.dept_id, section)
    subject_codes = [subject.subject_code for subject in subjects]
    students = Student.query.filter_by(section_id=section.section_id).order_by(Student.roll_no.asc()).all()

    payload = {
        'report_type': report_type,
        'section': {
            'id': section.section_id,
            'name': section.section_name,
            'batch_name': batch.batch_name if batch else None,
            'semester': section.current_semester,
        },
        'subjects': [
            {
                'code': subject.subject_code,
                'name': subject.subject_name,
            }
            for subject in subjects
        ],
        'rows': [],
    }

    if report_type == 'attendance':
        sessions = AttendanceSession.query.filter(
            AttendanceSession.section_id == section.section_id,
            AttendanceSession.subject_code.in_(subject_codes) if subject_codes else False,
        ).all()
        session_map = {session.session_id: session for session in sessions}
        session_ids = list(session_map.keys())
        total_by_subject = Counter(session.subject_code for session in sessions)
        records = AttendanceRecord.query.filter(AttendanceRecord.session_id.in_(session_ids)).all() if session_ids else []
        present_by_student_subject = Counter()
        for record in records:
            session = session_map.get(record.session_id)
            if session and record.status == 'P':
                present_by_student_subject[(record.student_id, session.subject_code)] += 1

        total_classes = sum(total_by_subject.values())
        for student in students:
            user = db.session.get(User, student.student_id)
            subject_values = {}
            attended_total = 0
            for code in subject_codes:
                present = present_by_student_subject[(student.student_id, code)]
                conducted = total_by_subject[code]
                attended_total += present
                subject_values[code] = {
                    'present': present,
                    'total': conducted,
                    'display': f'{present}/{conducted}',
                }
            percentage = round((attended_total / total_classes) * 100, 1) if total_classes else 0
            payload['rows'].append({
                'student_id': student.student_id,
                'roll_no': student.roll_no,
                'student_name': user.name if user else student.roll_no,
                'subjects': subject_values,
                'total_attended': attended_total,
                'total_classes': total_classes,
                'percentage': percentage,
            })
        return payload

    marks = Mark.query.filter(
        Mark.subject_code.in_(subject_codes) if subject_codes else False,
        Mark.student_id.in_([student.student_id for student in students]) if students else False,
    ).all()
    marks_by_student_subject = defaultdict(dict)
    for mark in marks:
        marks_by_student_subject[(mark.student_id, mark.subject_code)][mark.exam_type] = mark.obtained_marks

    for student in students:
        user = db.session.get(User, student.student_id)
        subject_values = {}
        for code in subject_codes:
            subject_total = calculate_subject_total(marks_by_student_subject[(student.student_id, code)])
            subject_values[code] = {
                'total': subject_total,
                'display': f'{subject_total}/30',
                'words': number_to_words(subject_total),
            }
        payload['rows'].append({
            'student_id': student.student_id,
            'roll_no': student.roll_no,
            'student_name': user.name if user else student.roll_no,
            'subjects': subject_values,
        })
    return payload


def class_report_headers(payload):
    base_headers = ['Roll No', 'Student Name']
    subject_headers = [subject['code'] for subject in payload['subjects']]
    if payload['report_type'] == 'attendance':
        return base_headers + subject_headers + ['Percentage']
    return base_headers + subject_headers


def class_report_row_values(payload, row):
    values = [row['roll_no'], row['student_name']]
    for subject in payload['subjects']:
        values.append(row['subjects'].get(subject['code'], {}).get('display', '0/0' if payload['report_type'] == 'attendance' else '0/30'))
    if payload['report_type'] == 'attendance':
        values.append(f"{row['percentage']}%")
    return values


def build_class_report_pdf(payload):
    page_width = 842
    page_height = 595
    margin = 28
    title = 'Class Attendance Report' if payload['report_type'] == 'attendance' else 'Class Marks Report'
    section = payload['section']
    subtitle = f"Section {section['name']} - {section.get('batch_name') or '-'} - Sem {section.get('semester') or '-'}"
    headers = class_report_headers(payload)
    rows = [class_report_row_values(payload, row) for row in payload['rows']]
    rows_per_page = 22
    pages = []
    subject_count = max(1, len(headers) - 3 if payload['report_type'] == 'attendance' else len(headers) - 2)

    fixed_width = 178 if payload['report_type'] == 'attendance' else 190
    percent_width = 62 if payload['report_type'] == 'attendance' else 0
    subject_width = max(45, (page_width - (margin * 2) - fixed_width - percent_width) / subject_count)
    widths = [62, fixed_width - 62] + [subject_width] * (len(payload['subjects']))
    if payload['report_type'] == 'attendance':
        widths.append(percent_width)

    for page_index in range(0, max(1, len(rows)), rows_per_page):
        page_rows = rows[page_index:page_index + rows_per_page]
        commands = [
            build_pdf_text(margin, page_height - 34, title, font='F2', size=13),
            build_pdf_text(margin, page_height - 52, subtitle, size=9),
        ]
        y = page_height - 82
        x = margin
        header_height = 24
        for header, width in zip(headers, widths):
            commands.append(build_pdf_rect(x, y - header_height, width, header_height, fill_gray=0.86))
            commands.append(build_pdf_text(x + 4, y - 15, truncate_text(header, max(6, int(width / 5))), font='F2', size=7))
            x += width

        row_height = 19
        y -= header_height
        for row_values in page_rows:
            x = margin
            y -= row_height
            for value, width in zip(row_values, widths):
                commands.append(build_pdf_rect(x, y, width, row_height, stroke_gray=0.45))
                commands.append(build_pdf_text(x + 4, y + 7, truncate_text(value, max(6, int(width / 5))), size=7))
                x += width
        commands.append(build_pdf_text(page_width - 92, margin - 2, f'Page {len(pages) + 1}', size=8))
        pages.append('\n'.join(commands))

    return build_pdf_pages(pages, page_width=page_width, page_height=page_height)


def build_class_report_xlsx(payload):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError('openpyxl is required for XLSX export') from exc

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance' if payload['report_type'] == 'attendance' else 'Marks'
    ws.append(class_report_headers(payload))
    for row in payload['rows']:
        ws.append(class_report_row_values(payload, row))
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 28)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@dept_bp.route('/class-reports/<int:section_id>/<report_type>', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_class_report(current_user, section_id, report_type):
    try:
        return jsonify(get_class_report_payload(current_user, section_id, report_type))
    except ValueError as exc:
        message = str(exc)
        return jsonify({'error': message}), 404 if 'not found' in message.lower() else 400


@dept_bp.route('/class-reports/<int:section_id>/<report_type>/download', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def download_class_report(current_user, section_id, report_type):
    file_format = (request.args.get('format') or 'pdf').lower()
    try:
        payload = get_class_report_payload(current_user, section_id, report_type)
    except ValueError as exc:
        message = str(exc)
        return jsonify({'error': message}), 404 if 'not found' in message.lower() else 400

    section = payload['section']
    filename_base = f"class-{section['name']}-{payload['report_type']}-report".replace(' ', '-').lower()
    if file_format == 'pdf':
        response = make_response(build_class_report_pdf(payload))
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={filename_base}.pdf'
        return response
    if file_format == 'xlsx':
        try:
            file_bytes = build_class_report_xlsx(payload)
        except RuntimeError as exc:
            return jsonify({'error': str(exc)}), 500
        response = make_response(file_bytes)
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename_base}.xlsx'
        return response
    return jsonify({'error': 'format must be pdf or xlsx'}), 400

@dept_bp.route('/sections/<int:section_id>/faculty', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_section_faculty(current_user, section_id):
    allocations = FacultyBatchSection.query.filter_by(section_id=section_id).all()
    faculty_ids = {a.faculty_id for a in allocations}
    data = []
    for fid in faculty_ids:
        faculty = db.session.get(Faculty, fid)
        user = db.session.get(User, faculty.user_id) if faculty else None
        data.append({'faculty_id': fid, 'name': user.name if user else None})
    return jsonify(data)

# ── Student Management ────────────────────────────────
@dept_bp.route('/students', methods=['GET'])
@token_required
@role_required('DEPT_ADMIN')
def get_students(current_user):
    query = Student.query.filter_by(dept_id=current_user.dept_id)

    batch_id = request.args.get('batch_id')
    program_id = request.args.get('program_id')
    semester = request.args.get('semester')
    attendance_lt = request.args.get('attendance_lt')
    student_type = request.args.get('type')

    if batch_id:
        try:
            query = query.filter_by(batch_id=int(batch_id))
        except ValueError:
            pass
    if program_id:
        try:
            b_ids = [b.batch_id for b in Batch.query.filter_by(program_id=int(program_id), dept_id=current_user.dept_id).all()]
            query = query.filter(Student.batch_id.in_(b_ids))
        except ValueError:
            pass
    if semester:
        try:
            sem = int(semester)
            section_ids = [s.section_id for s in Section.query.filter_by(current_semester=sem, dept_id=current_user.dept_id).all()]
            query = query.filter(Student.section_id.in_(section_ids))
        except ValueError:
            pass
    if student_type:
        query = query.filter_by(student_type=student_type)

    query = query.order_by(Student.roll_no.asc())
    students = query.all()
    student_ids = [s.student_id for s in students]
    section_ids_for_students = {s.section_id for s in students}
    batch_ids_for_students = {s.batch_id for s in students}

    users_by_id = {
        user.user_id: user
        for user in User.query.filter(User.user_id.in_(student_ids)).all()
    } if student_ids else {}
    missing_user_emails = [s.email for s in students if s.student_id not in users_by_id and s.email]
    users_by_email = {
        user.email.lower(): user
        for user in User.query.filter(User.email.in_(missing_user_emails)).all()
    } if missing_user_emails else {}
    sections_by_id = {
        section.section_id: section
        for section in Section.query.filter(Section.section_id.in_(section_ids_for_students)).all()
    } if section_ids_for_students else {}
    batches_by_id = {
        batch.batch_id: batch
        for batch in Batch.query.filter(Batch.batch_id.in_(batch_ids_for_students)).all()
    } if batch_ids_for_students else {}
    program_ids = {batch.program_id for batch in batches_by_id.values() if batch.program_id}
    programs_by_id = {
        program.program_id: program
        for program in Program.query.filter(Program.program_id.in_(program_ids)).all()
    } if program_ids else {}

    attendance_by_student = {}
    if student_ids:
        attendance_rows = (
            db.session.query(
                AttendanceRecord.student_id,
                func.count(AttendanceRecord.session_id),
                func.sum(case((AttendanceRecord.status == 'P', 1), else_=0)),
            )
            .filter(AttendanceRecord.student_id.in_(student_ids))
            .group_by(AttendanceRecord.student_id)
            .all()
        )
        attendance_by_student = {
            student_id: {'total': total or 0, 'present': present or 0}
            for student_id, total, present in attendance_rows
        }

    marks_by_student = defaultdict(dict)
    if student_ids:
        mark_rows = (
            db.session.query(Mark.student_id, Mark.exam_type, Mark.obtained_marks, Mark.max_marks)
            .filter(
                Mark.student_id.in_(student_ids),
                Mark.exam_type.in_(['mid1', 'mid2', 'assignment1', 'assignment2']),
            )
            .all()
        )
        for student_id, exam_type, obtained_marks, max_marks in mark_rows:
            marks_by_student[student_id][exam_type] = {
                'obtained_marks': obtained_marks or 0,
                'max_marks': max_marks or 0,
            }

    student_list = []
    for s in students:
        u = users_by_id.get(s.student_id) or users_by_email.get((s.email or '').lower())
        section = sections_by_id.get(s.section_id)
        batch = batches_by_id.get(s.batch_id)

        attendance = attendance_by_student.get(s.student_id, {'total': 0, 'present': 0})
        total = attendance['total']
        present = attendance['present']
        attendance_pct = round((present / total) * 100, 1) if total > 0 else 0

        student_marks = marks_by_student.get(s.student_id, {})
        mid1 = student_marks.get('mid1')
        mid2 = student_marks.get('mid2')
        assign1 = student_marks.get('assignment1')
        assign2 = student_marks.get('assignment2')

        best_mid = max(mid1['obtained_marks'] if mid1 else 0, mid2['obtained_marks'] if mid2 else 0)
        best_mid = min(best_mid, 20)
        assignment_total = 0
        if assign1:
            assignment_total += min(assign1['obtained_marks'], assign1['max_marks'])
        if assign2:
            assignment_total += min(assign2['obtained_marks'], assign2['max_marks'])
        assignment_total = min(assignment_total, 10)

        total_marks = round(best_mid + assignment_total, 2)

        if attendance_lt:
            try:
                if float(attendance_pct) >= float(attendance_lt):
                    continue
            except ValueError:
                pass

        student_list.append({
            'id': s.student_id,
            'roll_no': s.roll_no,
            'name': u.name if u else None,
            'email': s.email,
            'phone': s.phone,
            'student_type': s.student_type,
            'passport_number': s.passport_number,
            'category': s.category,
            'entrance_marks': s.entrance_marks,
            'section_id': s.section_id,
            'section_name': section.section_name if section else None,
            'batch_id': s.batch_id,
            'batch_name': batch.batch_name if batch else None,
            'program_id': batch.program_id if batch else None,
            'program_name': (programs_by_id.get(batch.program_id).program_name if batch and batch.program_id and programs_by_id.get(batch.program_id) else None),
            'attendance_pct': attendance_pct,
            'total_marks': total_marks,
        })

    return jsonify(student_list)

@dept_bp.route('/students', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def create_student(current_user):
    data = request.get_json()
    required = ['name', 'email', 'password', 'roll_no', 'batch_id', 'section_id', 'student_type', 'category']
    if not data or not all(data.get(k) for k in required):
        return jsonify({'error': 'name, email, password, roll_no, batch_id, section_id, student_type, category are required'}), 400
    # Ensure batch_id and section_id are integers
    try:
        batch_id = int(data['batch_id'])
        section_id = int(data['section_id'])
    except (ValueError, TypeError):
        return jsonify({'error': 'batch_id and section_id must be integers'}), 400
    # Check if batch and section are in the same dept
    batch = db.session.get(Batch, batch_id)
    section = db.session.get(Section, section_id)
    if not batch or batch.dept_id != current_user.dept_id:
        return jsonify({'error': 'Invalid batch'}), 400
    if not section or section.dept_id != current_user.dept_id:
        return jsonify({'error': 'Invalid section'}), 400
    if section.batch_id != batch_id:
        return jsonify({'error': 'Section does not belong to this batch'}), 400
    email = normalize_email(data['email'])
    roll_no = data['roll_no'].strip().upper()
    if User.query.filter_by(email=email).first():
        return jsonify({'error': 'Email already exists'}), 409
    if Student.query.filter_by(roll_no=roll_no, dept_id=current_user.dept_id).first():
        return jsonify({'error': 'Roll number already exists'}), 409
    phone = (data.get('phone') or '').strip()
    student_type = data.get('student_type', 'Regular')
    passport_number = (data.get('passport_number') or '').strip() or None
    category = (data.get('category') or '').strip()
    entrance_marks = float(data.get('entrance_marks') or 0)

    user = User(name=data['name'].strip(), email=email, password_hash=hash_password(data['password']), phone=phone, role='STUDENT', dept_id=current_user.dept_id, college_id=current_user.college_id)
    db.session.add(user)
    db.session.flush()
    student = Student(
        student_id=user.user_id,
        roll_no=roll_no,
        batch_id=batch_id,
        section_id=section_id,
        dept_id=current_user.dept_id,
        email=email,
        phone=phone,
        student_type=student_type,
        passport_number=passport_number,
        category=category,
        entrance_marks=entrance_marks,
    )
    db.session.add(student)
    db.session.commit()
    return jsonify({'id': student.student_id, 'name': user.name, 'roll_no': student.roll_no}), 201

@dept_bp.route('/students/<int:student_id>', methods=['PUT'])
@token_required
@role_required('DEPT_ADMIN')
def update_student(current_user, student_id):
    student = db.session.get(Student, student_id)
    if not student or student.dept_id != current_user.dept_id:
        return jsonify({'error': 'Student not found'}), 404

    data = request.get_json() or {}
    required = ['name', 'email', 'roll_no', 'batch_id', 'section_id', 'student_type', 'category']
    if not all(data.get(k) for k in required):
        return jsonify({'error': 'name, email, roll_no, batch_id, section_id, student_type, category are required'}), 400

    try:
        batch_id = int(data['batch_id'])
        section_id = int(data['section_id'])
    except (ValueError, TypeError):
        return jsonify({'error': 'batch_id and section_id must be integers'}), 400

    batch = db.session.get(Batch, batch_id)
    section = db.session.get(Section, section_id)
    if not batch or batch.dept_id != current_user.dept_id:
        return jsonify({'error': 'Invalid batch'}), 400
    if not section or section.dept_id != current_user.dept_id:
        return jsonify({'error': 'Invalid section'}), 400
    if section.batch_id != batch_id:
        return jsonify({'error': 'Section does not belong to this batch'}), 400

    email = normalize_email(data['email'])
    roll_no = data['roll_no'].strip().upper()
    existing_user = User.query.filter(User.email == email, User.user_id != student_id).first()
    if existing_user:
        return jsonify({'error': 'Email already exists'}), 409
    existing_student = Student.query.filter(
        Student.roll_no == roll_no,
        Student.dept_id == current_user.dept_id,
        Student.student_id != student_id,
    ).first()
    if existing_student:
        return jsonify({'error': 'Roll number already exists'}), 409

    user = db.session.get(User, student.student_id)
    phone = (data.get('phone') or '').strip()
    password = (data.get('password') or '').strip()

    if user:
        user.name = data['name'].strip()
        user.email = email
        user.phone = phone
        if password:
            user.password_hash = hash_password(password)

    student.roll_no = roll_no
    student.batch_id = batch_id
    student.section_id = section_id
    student.email = email
    student.phone = phone
    student.student_type = data.get('student_type', 'Regular')
    student.passport_number = (data.get('passport_number') or '').strip() or None
    student.category = (data.get('category') or '').strip()
    student.entrance_marks = float(data.get('entrance_marks') or 0)

    db.session.commit()
    return jsonify({'message': 'Student updated successfully'})


def parse_semester_value(value):
    if value is None or value == '':
        return 1
    text = str(value).strip().upper()
    if text.isdigit():
        return int(text)
    sem_prefix_match = re.search(r'\bSEM(?:ESTER)?\s*[-:]?\s*(\d+)\b', text)
    if sem_prefix_match:
        return int(sem_prefix_match.group(1))
    semester_match = re.search(r'(\d+)(?:ST|ND|RD|TH)?\s+SEMESTER', text)
    if semester_match:
        return int(semester_match.group(1))
    year_sem_match = re.match(r'^\s*(\d+)\s*-\s*(\d+)', text)
    if year_sem_match:
        year_no = int(year_sem_match.group(1))
        sem_in_year = int(year_sem_match.group(2))
        return ((year_no - 1) * 2) + sem_in_year
    roman_values = {'I': 1, 'II': 2, 'III': 3, 'IV': 4, 'V': 5, 'VI': 6, 'VII': 7, 'VIII': 8}
    return roman_values.get(text, 1)


def parse_batch_year_range(batch_name):
    match = re.match(r'^\s*(?P<start>\d{4})\s*-\s*(?P<end>\d{2}|\d{4})\s*$', str(batch_name or '').strip())
    if not match:
        return None
    start_year = int(match.group('start'))
    end_part = match.group('end')
    end_year = int(end_part) if len(end_part) == 4 else int(str(start_year)[:2] + end_part)
    return start_year, end_year


def parse_batch_year_span(batch_name):
    year_range = parse_batch_year_range(batch_name)
    if not year_range:
        return None
    start_year, end_year = year_range
    return end_year - start_year


def find_batch_by_identity(current_user, batch_identity, bulk_cache=None):
    batch_identity = clean_bulk_cell(batch_identity)
    if not batch_identity:
        return None

    if bulk_cache:
        direct = bulk_cache['batches_by_name'].get(batch_identity.lower())
        if direct:
            return direct
        year_range = parse_batch_year_range(batch_identity)
        if year_range:
            return bulk_cache['batches_by_year_range'].get(year_range)
        return None

    batch = Batch.query.filter_by(batch_name=batch_identity, dept_id=current_user.dept_id).first()
    if batch:
        return batch

    year_range = parse_batch_year_range(batch_identity)
    if not year_range:
        return None

    for dept_batch in Batch.query.filter_by(dept_id=current_user.dept_id).all():
        if parse_batch_year_range(dept_batch.batch_name) == year_range:
            return dept_batch
    return None


def clean_bulk_section_identity(value):
    text = clean_bulk_cell(value).upper()
    if not text:
        return ''
    text = re.sub(r'\([^)]*\bSEM(?:ESTER)?[^)]*\)', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\bSEM(?:ESTER)?\s*[-:]?\s*\d+\b', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\s+', ' ', text).strip(' -_')
    return text[:10] if text else ''


def find_section_by_identity(current_user, batch, section_identity, bulk_cache=None):
    section_identity = clean_bulk_section_identity(section_identity)
    if not section_identity:
        return None

    if bulk_cache:
        return bulk_cache['sections_by_batch_identity'].get((batch.batch_id, section_identity))

    sections = Section.query.filter_by(batch_id=batch.batch_id, dept_id=current_user.dept_id).all()
    for section in sections:
        if clean_bulk_section_identity(section.section_name) == section_identity:
            return section
    return None


def find_unique_batch_for_section(current_user, section_identity, bulk_cache=None):
    section_identity = clean_bulk_section_identity(section_identity)
    if not section_identity:
        return None

    if bulk_cache:
        batch_ids = bulk_cache['section_identity_batch_ids'].get(section_identity, set())
        if len(batch_ids) != 1:
            return None
        return bulk_cache['batches_by_id'].get(next(iter(batch_ids)))

    matches = []
    for section in Section.query.filter_by(dept_id=current_user.dept_id).all():
        if clean_bulk_section_identity(section.section_name) == section_identity:
            matches.append(section.batch_id)
    unique_batch_ids = set(matches)
    if len(unique_batch_ids) != 1:
        return None
    return db.session.get(Batch, next(iter(unique_batch_ids)))


def create_bulk_batch_if_possible(current_user, batch_name, current_semester, bulk_cache=None):
    dept_batches = bulk_cache['batches'] if bulk_cache else Batch.query.filter_by(dept_id=current_user.dept_id).all()
    program_ids = {batch.program_id for batch in dept_batches if batch.program_id}
    if len(program_ids) != 1:
        raise ValueError(f'Batch {batch_name} not found')

    program = db.session.get(Program, next(iter(program_ids)))
    year_span = parse_batch_year_span(batch_name)
    expected_span = (program.duration_semesters // 2) if program else None
    if not year_span or not expected_span or year_span != expected_span:
        raise ValueError(f'Batch {batch_name} not found')

    batch = Batch(batch_name=batch_name, dept_id=current_user.dept_id, program_id=program.program_id)
    db.session.add(batch)
    db.session.flush()
    if bulk_cache:
        bulk_cache['batches'].append(batch)
        bulk_cache['batches_by_id'][batch.batch_id] = batch
        bulk_cache['batches_by_name'][batch.batch_name.lower()] = batch
        year_range = parse_batch_year_range(batch.batch_name)
        if year_range:
            bulk_cache['batches_by_year_range'][year_range] = batch
    for semester_no in range(1, program.duration_semesters + 1):
        db.session.add(Semester(
            batch_id=batch.batch_id,
            semester_no=semester_no,
            is_active=(semester_no == current_semester),
        ))
    return batch


def resolve_bulk_student_target(current_user, row, bulk_cache=None):
    batch_id = clean_bulk_cell(row.get('batch_id'))
    section_id = clean_bulk_cell(row.get('section_id'))
    batch_name = clean_bulk_cell(row.get('batch_name'))
    batch_identity = batch_name or (batch_id if not batch_id.isdigit() else '')
    section_name = clean_bulk_cell(row.get('section_name') or row.get('class_name'))
    section_identity = section_name or (section_id if not section_id.isdigit() else '')
    current_semester = parse_semester_value(
        row.get('current_semester') or row.get('semester') or row.get('semester_text') or section_name or section_id
    )

    batch = None
    if batch_id and batch_id.isdigit():
        batch = bulk_cache['batches_by_id'].get(int(batch_id)) if bulk_cache else db.session.get(Batch, int(batch_id))
    elif batch_identity:
        batch = find_batch_by_identity(current_user, batch_identity, bulk_cache)
        if not batch:
            batch = create_bulk_batch_if_possible(current_user, batch_identity, current_semester, bulk_cache)
    else:
        batch = find_unique_batch_for_section(current_user, section_identity, bulk_cache)
        if not batch:
            dept_batches = bulk_cache['batches'] if bulk_cache else Batch.query.filter_by(dept_id=current_user.dept_id).all()
            if len(dept_batches) == 1:
                batch = dept_batches[0]

    if not batch or batch.dept_id != current_user.dept_id:
        raise ValueError('Invalid or missing batch')

    section = None
    if section_id and section_id.isdigit():
        section = bulk_cache['sections_by_id'].get(int(section_id)) if bulk_cache else db.session.get(Section, int(section_id))
    elif section_identity:
        section = find_section_by_identity(current_user, batch, section_identity, bulk_cache)
        if not section:
            cleaned_section_name = clean_bulk_section_identity(section_identity)
            section = Section(
                section_name=cleaned_section_name,
                current_semester=current_semester,
                batch_id=batch.batch_id,
                dept_id=current_user.dept_id,
                program_id=batch.program_id,
            )
            db.session.add(section)
            db.session.flush()
            if bulk_cache:
                bulk_cache['sections'].append(section)
                bulk_cache['sections_by_id'][section.section_id] = section
                bulk_cache['sections_by_batch_id'].setdefault(batch.batch_id, []).append(section)
                bulk_cache['sections_by_batch_identity'][(batch.batch_id, cleaned_section_name)] = section
                bulk_cache['section_identity_batch_ids'].setdefault(cleaned_section_name, set()).add(batch.batch_id)
    else:
        sections = bulk_cache['sections_by_batch_id'].get(batch.batch_id, []) if bulk_cache else Section.query.filter_by(batch_id=batch.batch_id, dept_id=current_user.dept_id).all()
        if len(sections) == 1:
            section = sections[0]
        elif not sections:
            section = Section(
                section_name='A',
                current_semester=current_semester,
                batch_id=batch.batch_id,
                dept_id=current_user.dept_id,
                program_id=batch.program_id,
            )
            db.session.add(section)
            db.session.flush()
            if bulk_cache:
                bulk_cache['sections'].append(section)
                bulk_cache['sections_by_id'][section.section_id] = section
                bulk_cache['sections_by_batch_id'].setdefault(batch.batch_id, []).append(section)
                bulk_cache['sections_by_batch_identity'][(batch.batch_id, 'A')] = section
                bulk_cache['section_identity_batch_ids'].setdefault('A', set()).add(batch.batch_id)

    if not section or section.dept_id != current_user.dept_id or section.batch_id != batch.batch_id:
        raise ValueError('Invalid or missing section')

    return batch, section


def build_bulk_email(email, roll_no, current_user, idx):
    email = normalize_email(clean_bulk_cell(email))
    if email:
        return email
    seed = clean_bulk_cell(roll_no, f'student{idx}').lower()
    seed = re.sub(r'[^a-z0-9]+', '.', seed).strip('.') or f'student{idx}'
    return f'{seed}.{current_user.dept_id}.{idx}@attendance.local'


def safe_bulk_float(value, default=0):
    try:
        return float(clean_bulk_cell(value, str(default)))
    except (TypeError, ValueError):
        return default


def build_bulk_upload_cache(current_user):
    batches = Batch.query.filter_by(dept_id=current_user.dept_id).all()
    sections = Section.query.filter_by(dept_id=current_user.dept_id).all()

    cache = {
        'batches': batches,
        'batches_by_id': {batch.batch_id: batch for batch in batches},
        'batches_by_name': {batch.batch_name.lower(): batch for batch in batches},
        'batches_by_year_range': {},
        'sections': sections,
        'sections_by_id': {section.section_id: section for section in sections},
        'sections_by_batch_id': defaultdict(list),
        'sections_by_batch_identity': {},
        'section_identity_batch_ids': defaultdict(set),
    }

    for batch in batches:
        year_range = parse_batch_year_range(batch.batch_name)
        if year_range:
            cache['batches_by_year_range'][year_range] = batch

    for section in sections:
        section_identity = clean_bulk_section_identity(section.section_name)
        cache['sections_by_batch_id'][section.batch_id].append(section)
        if section_identity:
            cache['sections_by_batch_identity'][(section.batch_id, section_identity)] = section
            cache['section_identity_batch_ids'][section_identity].add(section.batch_id)

    return cache


def clean_bulk_cell(value, default=''):
    if value is None:
        return default
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text if text else default


def parse_bulk_rows(upload_file, filename):
    rows = []
    if filename.endswith('.csv'):
        import csv
        decoded = upload_file.stream.read().decode('utf-8', errors='replace')
        reader = csv.DictReader(decoded.splitlines())
        for row in reader:
            rows.append({normalize_bulk_header(key): value for key, value in row.items() if key})
        return rows

    from openpyxl import load_workbook
    wb = load_workbook(upload_file, data_only=True)
    sheet = wb.active
    raw_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    headers = [normalize_bulk_header(header) for header in raw_headers]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
        if any(value not in (None, '') for value in row_data.values()):
            rows.append(row_data)
    return rows


@dept_bp.route('/students/bulk', methods=['POST'])
@token_required
@role_required('DEPT_ADMIN')
def bulk_upload_students(current_user):
    csv_file = request.files.get('file')
    if not csv_file:
        return jsonify({'error': 'file is required'}), 400

    filename = csv_file.filename.lower()
    supported = ['.csv', '.xlsx', '.xls']
    if not any(filename.endswith(ext) for ext in supported):
        return jsonify({'error': 'Only CSV/XLSX files are supported'}), 400

    try:
        if filename.endswith('.csv'):
            rows = parse_bulk_rows(csv_file, filename)
        else:
            rows = parse_bulk_rows(csv_file, filename)
    except ImportError:
        return jsonify({'error': 'openpyxl is required for Excel upload'}), 500
    except Exception as exc:
        if filename.endswith('.csv'):
            return jsonify({'error': f'CSV parse error: {exc}'}), 400
        return jsonify({'error': f'Excel parse error: {exc}'}), 400

    bulk_cache = build_bulk_upload_cache(current_user)
    upload_emails = set()
    upload_rolls = set()
    for idx, row in enumerate(rows, start=1):
        row_roll = clean_bulk_cell(get_bulk_value(row, 'roll_no', 'roll no', 'roll', 'roll number', 'enrollment_no', 'enrollment no'))
        row_email = build_bulk_email(get_bulk_value(row, 'email', 'student_email', 'student email'), row_roll, current_user, idx)
        if not row_roll:
            row_roll = re.sub(r'[^A-Z0-9]+', '', row_email.split('@')[0].upper()) or f'BULK{idx}'
        upload_emails.add(normalize_email(row_email))
        upload_rolls.add(row_roll.upper())

    existing_emails = {
        email.lower()
        for (email,) in db.session.query(User.email).filter(User.email.in_(upload_emails)).all()
        if email
    } if upload_emails else set()
    existing_rolls = {
        roll_no.upper()
        for (roll_no,) in db.session.query(Student.roll_no)
        .filter(Student.dept_id == current_user.dept_id, Student.roll_no.in_(upload_rolls))
        .all()
        if roll_no
    } if upload_rolls else set()

    created = 0
    errors = []
    new_emails = set()
    new_rolls = set()
    password_hash_cache = {}
    pending_emails = []
    pending_rolls = []
    pending_created = 0
    for idx, row in enumerate(rows, start=1):
        try:
            name = get_bulk_value(row, 'name', 'student_name', 'student name')
            email = get_bulk_value(row, 'email', 'student_email', 'student email')
            roll_no = get_bulk_value(row, 'roll_no', 'roll no', 'roll', 'roll number', 'enrollment_no', 'enrollment no')
            clean_roll_no = clean_bulk_cell(roll_no)
            clean_email = build_bulk_email(email, clean_roll_no, current_user, idx)
            if not clean_roll_no:
                clean_roll_no = re.sub(r'[^A-Z0-9]+', '', clean_email.split('@')[0].upper()) or f'BULK{idx}'
            data = {
                'name': clean_bulk_cell(name, clean_roll_no),
                'email': clean_email,
                'password': clean_bulk_cell(get_bulk_value(row, 'password'), 'Student@123'),
                'roll_no': clean_roll_no,
                'batch_id': get_bulk_value(row, 'batch_id', 'batch id'),
                'section_id': get_bulk_value(row, 'section_id', 'section id'),
                'batch_name': get_bulk_value(row, 'batch_name', 'batch name', 'batch'),
                'section_name': get_bulk_value(row, 'section_name', 'section name', 'section', 'class_name', 'class name'),
                'current_semester': get_bulk_value(row, 'current_semester', 'current semester', 'semester', 'semester_text', 'semester text'),
                'student_type': clean_bulk_cell(get_bulk_value(row, 'student_type', 'student type', 'type'), 'Regular')[:20],
                'passport_number': clean_bulk_cell(get_bulk_value(row, 'passport_number', 'passport number', 'passport'))[:120],
                'category': clean_bulk_cell(get_bulk_value(row, 'category'))[:120],
                'entrance_marks': get_bulk_value(row, 'entrance_marks', 'entrance marks') or 0,
                'phone': clean_bulk_cell(get_bulk_value(row, 'phone', 'mobile', 'phone number', 'mobile number'))[:20],
            }

            try:
                batch, section = resolve_bulk_student_target(current_user, data, bulk_cache)
            except ValueError as exc:
                errors.append({'row': idx, 'error': str(exc)})
                continue

            email = normalize_email(data['email'])
            roll_no = data['roll_no'].upper()
            if email in existing_emails or email in new_emails or roll_no in existing_rolls or roll_no in new_rolls:
                errors.append({'row': idx, 'error': 'Duplicate user or roll number'});
                continue

            password_hash = password_hash_cache.get(data['password'])
            if not password_hash:
                password_hash = hash_password(data['password'])
                password_hash_cache[data['password']] = password_hash

            user = User(name=data['name'], email=email, password_hash=password_hash, phone=data['phone'], role='STUDENT', dept_id=current_user.dept_id, college_id=current_user.college_id)
            db.session.add(user)
            db.session.flush()
            student = Student(
                student_id=user.user_id,
                roll_no=roll_no,
                batch_id=batch.batch_id,
                section_id=section.section_id,
                dept_id=current_user.dept_id,
                email=email,
                phone=data['phone'],
                student_type=data['student_type'],
                passport_number=data['passport_number'] or None,
                category=data['category'] or None,
                entrance_marks=safe_bulk_float(data['entrance_marks']),
            )
            db.session.add(student)
            new_emails.add(email)
            new_rolls.add(roll_no)
            pending_emails.append(email)
            pending_rolls.append(roll_no)
            pending_created += 1
            created += 1
            if pending_created >= 100:
                db.session.commit()
                pending_emails.clear()
                pending_rolls.clear()
                pending_created = 0
        except Exception as exc:
            db.session.rollback()
            for pending_email in pending_emails:
                new_emails.discard(pending_email)
            for pending_roll in pending_rolls:
                new_rolls.discard(pending_roll)
            created -= pending_created
            pending_emails.clear()
            pending_rolls.clear()
            pending_created = 0
            bulk_cache = build_bulk_upload_cache(current_user)
            errors.append({'row': idx, 'error': exc.__class__.__name__})
            continue

    if pending_created:
        db.session.commit()
    return jsonify({'created': created, 'errors': errors})

@dept_bp.route('/students/<int:student_id>', methods=['DELETE'])
@token_required
@role_required('DEPT_ADMIN')
def delete_student(current_user, student_id):
    student = db.session.get(Student, student_id)
    if not student or student.dept_id != current_user.dept_id:
        return jsonify({'error': 'Student not found'}), 404
    user = db.session.get(User, student.student_id)
    db.session.delete(student)
    if user:
        db.session.delete(user)
    db.session.commit()
    return jsonify({'message': 'Student deleted'})

@dept_bp.route('/batches/<int:batch_id>/students', methods=['DELETE'])
@token_required
@role_required('DEPT_ADMIN')
def delete_batch_students(current_user, batch_id):
    """Delete all students in a batch"""
    batch = db.session.get(Batch, batch_id)
    if not batch or batch.dept_id != current_user.dept_id:
        return jsonify({'error': 'Batch not found'}), 404
    students = Student.query.filter_by(batch_id=batch_id, dept_id=current_user.dept_id).all()
    count = len(students)
    for student in students:
        user = db.session.get(User, student.student_id)
        db.session.delete(student)
        if user:
            db.session.delete(user)
    db.session.commit()
    return jsonify({'message': f'Deleted {count} students from batch', 'count': count})
